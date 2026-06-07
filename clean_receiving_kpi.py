"""KPI #4: 入库单 KPI 清洗与统计

基于入库单 Excel 导出数据，统计：
  - 收货总量/总件数/总SKU
  - 供应商维度统计
  - 收货效率（行/小时）
  - 上架及时率
  - 入库时效分析

用法:
  python clean_receiving_kpi.py receiving_20260608.xlsx --kpi-date 2026-06-08
"""
from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Sheet 名称 ──────────────────────────────────────────────
RECEIVING_SUMMARY = "入库汇总"
SUPPLIER_SUMMARY = "供应商统计"
RECEIVING_EFFICIENCY = "收货效率"
RECEIVING_LEADTIME = "入库时效"

# ── 候选列名（中英文双语，兼容各种导出模板）─────────────────
ASN_NO_CANDIDATES = [
    "ASN No./入库单号", "ASN No.", "入库单号", "Receiving No./收货单号",
    "Receiving No.", "收货单号", "asnNo", "receivingNo", "orderNo",
]

SUPPLIER_CANDIDATES = [
    "Supplier/供应商", "Supplier", "供应商", "Supplier Name/供应商名称",
    "Supplier Name", "供应商名称", "supplierName", "supplier",
]

WH_CODE_CANDIDATES = [
    "WH Code/仓库编码", "WH Code", "仓库编码", "Warehouse/仓库",
    "Warehouse", "仓库", "whCode", "whName",
]

SKU_CODE_CANDIDATES = [
    "SKU/产品编码", "SKU", "产品编码", "skuCode",
    "Seller SKU/商家SKU", "Seller SKU", "商家SKU",
    "Product Code/产品代码", "Product Code", "产品代码",
]

QTY_CANDIDATES = [
    "Qty/数量", "Qty", "数量", "Receiving Qty/收货数量",
    "Receiving Qty", "收货数量", "qty", "quantity",
    "Total Qty/总数量", "Total Qty", "总数量",
]

RECEIVE_TIME_CANDIDATES = [
    "Receive time/收货时间", "Receive time", "收货时间",
    "Receiving time/收货时间", "Receiving time",
    "receiveTime", "receivingTime", "Complete time/完成时间",
    "Complete time", "完成时间",
]

CREATE_TIME_CANDIDATES = [
    "Create time/创建时间", "Create time", "创建时间", "createTime",
    "ASN time/ASN时间", "ASN time", "ASN时间",
]

PUTAWAY_TIME_CANDIDATES = [
    "Putaway time/上架时间", "Putaway time", "上架时间",
    "putawayTime", "Shelve time/上架时间", "Shelve time",
]

STATUS_CANDIDATES = [
    "Status/状态", "Status", "状态", "status",
]

OPERATOR_CANDIDATES = [
    "Operator/操作人", "Operator", "操作人", "Receiver/收货人",
    "Receiver", "收货人", "operatorName", "receiverName",
]

# ── 表头 ────────────────────────────────────────────────────
SUMMARY_HEADERS = [
    "日期", "周", "月",
    "入库单数", "供应商数", "SKU数", "总数量",
    "已完结单数", "进行中单数", "完结率",
]

SUPPLIER_HEADERS = [
    "供应商", "入库单数", "SKU数", "总数量",
    "最早收货", "最晚收货", "平均时效(h)", "完结率",
]

EFFICIENCY_HEADERS = [
    "日期", "收货人", "入库单数", "SKU数", "总数量",
    "最早收货", "最晚收货", "工作时长(h)", "效率(件/小时)", "效率(行/小时)",
]

LEADTIME_HEADERS = [
    "时效区间", "入库单数", "占比",
]

# ── 样式 ────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GREEN_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
RED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")


# ── 辅助函数 ────────────────────────────────────────────────
def compact_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def find_header(headers: list[Any], candidates: list[str]) -> int:
    normalized = [compact_header(c) for c in candidates]
    for index, header in enumerate(headers):
        if compact_header(header) in normalized:
            return index
    raise RuntimeError(f"Excel missing required column. Tried: {candidates}")


def find_header_optional(headers: list[Any], candidates: list[str]) -> int | None:
    try:
        return find_header(headers, candidates)
    except RuntimeError:
        return None


def parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def format_date(value: date) -> str:
    return value.strftime("%Y/%m/%d")


def format_datetime(value: datetime | None) -> str:
    return "" if value is None else value.strftime("%Y/%m/%d %H:%M:%S")


def round2(value: float) -> float:
    return round(value, 2)


def median(values: list[float]) -> float:
    if not values:
        return 0.0
    sorted_vals = sorted(values)
    n = len(sorted_vals)
    if n % 2 == 0:
        return round2((sorted_vals[n // 2 - 1] + sorted_vals[n // 2]) / 2)
    else:
        return round2(sorted_vals[n // 2])


# ── 样式设置 ────────────────────────────────────────────────
def style_sheet(ws: Any) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_width(ws: Any, max_width: int = 36) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, max_width))
        ws.column_dimensions[column_letter].width = width


# ── 数据提取 ────────────────────────────────────────────────
def extract_receiving_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    """从入库单 sheet 提取数据"""
    if sheet_name not in workbook.sheetnames:
        # 尝试找第一个 sheet
        if workbook.sheetnames:
            sheet_name = workbook.sheetnames[0]
        else:
            raise RuntimeError("Excel has no sheets")

    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]

    print(f"Processing sheet: {sheet_name}")
    print(f"Headers found ({len(headers)}): {[compact_header(h) for h in headers[:15]]}...")

    asn_idx = find_header_optional(headers, ASN_NO_CANDIDATES)
    supplier_idx = find_header_optional(headers, SUPPLIER_CANDIDATES)
    wh_idx = find_header_optional(headers, WH_CODE_CANDIDATES)
    sku_idx = find_header_optional(headers, SKU_CODE_CANDIDATES)
    qty_idx = find_header_optional(headers, QTY_CANDIDATES)
    receive_time_idx = find_header_optional(headers, RECEIVE_TIME_CANDIDATES)
    create_time_idx = find_header_optional(headers, CREATE_TIME_CANDIDATES)
    putaway_time_idx = find_header_optional(headers, PUTAWAY_TIME_CANDIDATES)
    status_idx = find_header_optional(headers, STATUS_CANDIDATES)
    operator_idx = find_header_optional(headers, OPERATOR_CANDIDATES)

    # 至少需要一个唯一标识列
    if asn_idx is None and sku_idx is None:
        print("WARNING: Could not find ASN/order number or SKU column. Using row index as identifier.")

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        data = {
            "_row": row_idx,
            "asn_no": str(row[asn_idx] or "").strip() if asn_idx is not None else f"ROW_{row_idx}",
            "supplier": str(row[supplier_idx] or "").strip() if supplier_idx is not None else "",
            "wh_code": str(row[wh_idx] or "").strip() if wh_idx is not None else "",
            "sku_code": str(row[sku_idx] or "").strip() if sku_idx is not None else "",
            "qty": parse_number(row[qty_idx]) if qty_idx is not None else None,
            "receive_time": parse_datetime(row[receive_time_idx]) if receive_time_idx is not None else None,
            "create_time": parse_datetime(row[create_time_idx]) if create_time_idx is not None else None,
            "putaway_time": parse_datetime(row[putaway_time_idx]) if putaway_time_idx is not None else None,
            "status": str(row[status_idx] or "").strip() if status_idx is not None else "",
            "operator": str(row[operator_idx] or "").strip() if operator_idx is not None else "",
        }
        rows.append(data)

    return rows


# ── 入库汇总 ────────────────────────────────────────────────
def build_receiving_summary(
    workbook: Any,
    rows: list[dict[str, Any]],
    kpi_day: date,
) -> dict[str, Any]:
    """生成入库汇总统计"""
    iso_year, iso_week, _ = kpi_day.isocalendar()

    asn_set: set[str] = set()
    supplier_set: set[str] = set()
    sku_set: set[str] = set()
    total_qty = 0.0
    completed = 0
    in_progress = 0

    for row in rows:
        if row["asn_no"]:
            asn_set.add(row["asn_no"])
        if row["supplier"]:
            supplier_set.add(row["supplier"])
        if row["sku_code"]:
            sku_set.add(row["sku_code"])
        if row["qty"] is not None:
            total_qty += row["qty"]

        status_lower = row["status"].lower()
        if any(kw in status_lower for kw in ["完成", "完结", "closed", "done", "complete", "finished"]):
            completed += 1
        elif any(kw in status_lower for kw in ["进行", "处理", "in progress", "processing", "pending"]):
            in_progress += 1

    total = len(rows)
    completion_rate = round2(completed / total * 100) if total > 0 else 0

    qty_display = int(total_qty) if total_qty and total_qty.is_integer() else round2(total_qty)

    summary = {
        "日期": format_date(kpi_day),
        "周": f"{iso_year}-W{iso_week:02d}",
        "月": kpi_day.replace(day=1).strftime("%Y/%m/%d"),
        "入库单数": len(asn_set),
        "供应商数": len(supplier_set),
        "SKU数": len(sku_set),
        "总数量": qty_display,
        "已完结单数": completed,
        "进行中单数": in_progress,
        "完结率": completion_rate,
    }

    if RECEIVING_SUMMARY in workbook.sheetnames:
        del workbook[RECEIVING_SUMMARY]
    ws = workbook.create_sheet(RECEIVING_SUMMARY, 0)
    ws.append(SUMMARY_HEADERS)
    ws.append([summary[h] for h in SUMMARY_HEADERS])
    style_sheet(ws)
    auto_width(ws)

    return summary


# ── 供应商统计 ──────────────────────────────────────────────
def build_supplier_summary(workbook: Any, rows: list[dict[str, Any]]) -> None:
    """按供应商聚合统计"""
    supplier_data: dict[str, dict[str, Any]] = {}
    for row in rows:
        supplier = row["supplier"] or "(未知供应商)"
        if supplier not in supplier_data:
            supplier_data[supplier] = {
                "asns": set(), "skus": set(), "qty": 0.0,
                "min_time": None, "max_time": None,
                "leadtimes": [],
                "completed": 0, "total": 0,
            }

        sd = supplier_data[supplier]
        if row["asn_no"]:
            sd["asns"].add(row["asn_no"])
        if row["sku_code"]:
            sd["skus"].add(row["sku_code"])
        if row["qty"] is not None:
            sd["qty"] += row["qty"]

        rt = row["receive_time"] or row["create_time"]
        if rt is not None:
            sd["min_time"] = rt if sd["min_time"] is None else min(sd["min_time"], rt)
            sd["max_time"] = rt if sd["max_time"] is None else max(sd["max_time"], rt)

        if row["create_time"] and row["receive_time"]:
            lt = (row["receive_time"] - row["create_time"]).total_seconds() / 3600.0
            if lt >= 0:
                sd["leadtimes"].append(lt)

        sd["total"] += 1
        status_lower = row["status"].lower()
        if any(kw in status_lower for kw in ["完成", "完结", "closed", "done", "complete"]):
            sd["completed"] += 1

    if SUPPLIER_SUMMARY in workbook.sheetnames:
        del workbook[SUPPLIER_SUMMARY]
    ws = workbook.create_sheet(SUPPLIER_SUMMARY)
    ws.append(SUPPLIER_HEADERS)

    for supplier, sd in sorted(supplier_data.items(), key=lambda x: x[1]["qty"], reverse=True):
        qty = int(sd["qty"]) if sd["qty"].is_integer() else round2(sd["qty"])
        avg_lt = round2(sum(sd["leadtimes"]) / len(sd["leadtimes"])) if sd["leadtimes"] else ""
        completion = round2(sd["completed"] / sd["total"] * 100) if sd["total"] > 0 else 0

        ws.append([
            supplier, len(sd["asns"]), len(sd["skus"]), qty,
            format_datetime(sd["min_time"]), format_datetime(sd["max_time"]),
            avg_lt, completion,
        ])

    style_sheet(ws)
    auto_width(ws)


# ── 收货效率 ────────────────────────────────────────────────
def build_efficiency_sheet(workbook: Any, rows: list[dict[str, Any]]) -> None:
    """按收货人统计收货效率"""
    operator_data: dict[str, dict[str, Any]] = {}
    for row in rows:
        operator = row["operator"] or "(未知收货人)"
        if operator not in operator_data:
            operator_data[operator] = {
                "asns": set(), "skus": set(), "qty": 0.0,
                "min_time": None, "max_time": None,
                "rows_count": 0,
            }

        od = operator_data[operator]
        if row["asn_no"]:
            od["asns"].add(row["asn_no"])
        if row["sku_code"]:
            od["skus"].add(row["sku_code"])
        if row["qty"] is not None:
            od["qty"] += row["qty"]
        od["rows_count"] += 1

        rt = row["receive_time"] or row["create_time"]
        if rt is not None:
            od["min_time"] = rt if od["min_time"] is None else min(od["min_time"], rt)
            od["max_time"] = rt if od["max_time"] is None else max(od["max_time"], rt)

    if RECEIVING_EFFICIENCY in workbook.sheetnames:
        del workbook[RECEIVING_EFFICIENCY]
    ws = workbook.create_sheet(RECEIVING_EFFICIENCY)
    ws.append(EFFICIENCY_HEADERS)

    today_str = format_date(date.today())

    for operator, od in sorted(operator_data.items(), key=lambda x: x[1]["qty"], reverse=True):
        qty = int(od["qty"]) if od["qty"].is_integer() else round2(od["qty"])

        work_hours = 0.0
        if od["min_time"] and od["max_time"]:
            work_hours = round2((od["max_time"] - od["min_time"]).total_seconds() / 3600.0)

        eff_per_hour = round2(qty / work_hours) if work_hours > 0 else 0.0
        lines_per_hour = round2(od["rows_count"] / work_hours) if work_hours > 0 else 0.0

        ws.append([
            today_str, operator, len(od["asns"]), len(od["skus"]), qty,
            format_datetime(od["min_time"]), format_datetime(od["max_time"]),
            work_hours, eff_per_hour, lines_per_hour,
        ])

    style_sheet(ws)
    auto_width(ws)


# ── 入库时效分布 ────────────────────────────────────────────
def build_leadtime_sheet(workbook: Any, rows: list[dict[str, Any]]) -> None:
    """生成入库时效分布"""
    buckets = [
        ("≤1h", 0, 1),
        ("1-2h", 1, 2),
        ("2-4h", 2, 4),
        ("4-8h", 4, 8),
        ("8-24h", 8, 24),
        ("24-48h", 24, 48),
        (">48h", 48, float("inf")),
    ]

    leadtimes = []
    for row in rows:
        if row["create_time"] and row["receive_time"]:
            lt = (row["receive_time"] - row["create_time"]).total_seconds() / 3600.0
            if lt >= 0:
                leadtimes.append(lt)

    total = len(leadtimes) if leadtimes else 1

    if RECEIVING_LEADTIME in workbook.sheetnames:
        del workbook[RECEIVING_LEADTIME]
    ws = workbook.create_sheet(RECEIVING_LEADTIME)
    ws.append(LEADTIME_HEADERS)

    for label, low, high in buckets:
        count = sum(1 for lt in leadtimes if low < lt <= high)
        pct = round2(count / total * 100)
        ws.append([label, count, f"{pct}%"])

    style_sheet(ws)
    auto_width(ws)


# ── 主流程 ──────────────────────────────────────────────────
def clean_receiving(
    input_path: Path,
    output_path: Path,
    kpi_day: date,
    sheet_name: str = "",
) -> None:
    workbook = load_workbook(input_path)

    # 自动检测 sheet（入库单通常命名为"入库单"或第一个 sheet）
    if not sheet_name:
        for candidate in ["入库单", "收货单", "ASN", "Receiving", "入库明细"]:
            if candidate in workbook.sheetnames:
                sheet_name = candidate
                break
        if not sheet_name and workbook.sheetnames:
            sheet_name = workbook.sheetnames[0]

    if not sheet_name or sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"Cannot find a valid sheet. Available: {workbook.sheetnames}")

    rows = extract_receiving_rows(workbook, sheet_name)
    print(f"Extracted {len(rows)} rows from sheet '{sheet_name}'")

    summary = build_receiving_summary(workbook, rows, kpi_day)
    build_supplier_summary(workbook, rows)
    build_efficiency_sheet(workbook, rows)
    build_leadtime_sheet(workbook, rows)

    # 删除不需要的 sheet，保留原始数据和生成的 KPI sheet
    target_sheets = {RECEIVING_SUMMARY, SUPPLIER_SUMMARY, RECEIVING_EFFICIENCY, RECEIVING_LEADTIME, sheet_name}
    for name in list(workbook.sheetnames):
        if name not in target_sheets:
            del workbook[name]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"\n{'='*60}")
    print(f"KPI date: {kpi_day:%Y-%m-%d}")
    print(f"Receiving Summary:")
    for k, v in summary.items():
        print(f"  {k}: {v}")
    print(f"\nOutput: {output_path.resolve()}")
    print(f"{'='*60}")


def main() -> None:
    parser = argparse.ArgumentParser(description="入库单 KPI 清洗与统计")
    parser.add_argument("input", help="Raw receiving export Excel path")
    parser.add_argument("--kpi-date", required=True, help="KPI date, YYYY-MM-DD")
    parser.add_argument("--sheet-name", default="", help="Target sheet name (auto-detect if empty)")
    parser.add_argument("--output", default="", help="Output Excel path")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if args.output:
        output_path = Path(args.output).resolve()
    else:
        output_dir = input_path.parent / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{input_path.stem}_receiving_kpi.xlsx"

    kpi_day = datetime.strptime(args.kpi_date, "%Y-%m-%d").date()
    clean_receiving(input_path, output_path, kpi_day, args.sheet_name)


if __name__ == "__main__":
    main()
