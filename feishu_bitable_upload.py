from __future__ import annotations

import argparse
import os
import re
import time
from dataclasses import dataclass
from datetime import date, datetime, time as datetime_time
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

import requests
from openpyxl import load_workbook


OPEN_FEISHU_API = "https://open.feishu.cn/open-apis"

DEFAULT_SHEET_NAME = "波次号去重统计"
DATE_HEADERS = {"日期", "月", "最早拣货时间", "最晚拣货时间"}

FIELD_TYPE_TEXT = 1
FIELD_TYPE_NUMBER = 2
FIELD_TYPE_DATE = 5


@dataclass
class UploadConfig:
    app_id: str
    app_secret: str
    app_token: str
    table_id: str
    input_path: Path
    sheet_name: str = DEFAULT_SHEET_NAME
    batch_size: int = 500
    create_missing_fields: bool = True
    clear_table: bool = False
    dry_run: bool = False
    request_interval: float = 0.2


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def require_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing environment variable: {name}")
    return value


def parse_bitable_url(url: str) -> tuple[str, str | None]:
    parsed = urlparse(url.strip())
    if "/wiki/" in parsed.path:
        raise RuntimeError(
            "This is a wiki URL. Please pass FEISHU_APP_TOKEN and FEISHU_TABLE_ID, "
            "or copy the direct /base/ URL from the Bitable page."
        )

    match = re.search(r"/base/([^/?#]+)", parsed.path)
    if not match:
        raise RuntimeError("Could not parse app_token from Bitable URL. Expected a /base/<app_token> URL.")

    app_token = match.group(1)
    query = parse_qs(parsed.query)
    table_id = (query.get("table") or query.get("table_id") or [None])[0]
    return app_token, table_id


def normalize_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def parse_datetime_value(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime_time.min)

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def datetime_to_millis(value: datetime) -> int:
    return int(value.timestamp() * 1000)


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def infer_field_type(header: str, values: list[Any]) -> int:
    non_empty = [value for value in values if value not in (None, "")]
    if header in DATE_HEADERS:
        return FIELD_TYPE_DATE
    if non_empty and all(is_number(value) for value in non_empty):
        return FIELD_TYPE_NUMBER
    return FIELD_TYPE_TEXT


def read_sheet(input_path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]], dict[str, int]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"Sheet not found: {sheet_name}. Available sheets: {', '.join(workbook.sheetnames)}")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise RuntimeError(f"Sheet is empty: {sheet_name}") from exc

    headers = [normalize_header(value) for value in raw_headers]
    headers = [header for header in headers if header]
    if not headers:
        raise RuntimeError(f"Sheet has no headers: {sheet_name}")
    if len(headers) != len(set(headers)):
        raise RuntimeError(f"Sheet has duplicate headers: {sheet_name}")

    records: list[dict[str, Any]] = []
    columns: dict[str, list[Any]] = {header: [] for header in headers}
    for row in rows:
        row_values = list(row[: len(headers)])
        if not any(value not in (None, "") for value in row_values):
            continue

        record: dict[str, Any] = {}
        for header, value in zip(headers, row_values):
            record[header] = value
            columns[header].append(value)
        records.append(record)

    inferred_types = {header: infer_field_type(header, values) for header, values in columns.items()}
    return headers, records, inferred_types


def feishu_request(
    method: str,
    path: str,
    *,
    token: str | None = None,
    json_body: dict[str, Any] | None = None,
    params: dict[str, Any] | None = None,
) -> dict[str, Any]:
    headers = {"Content-Type": "application/json; charset=utf-8"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    response = requests.request(
        method,
        f"{OPEN_FEISHU_API}{path}",
        headers=headers,
        json=json_body,
        params=params,
        timeout=60,
    )
    response.raise_for_status()
    payload = response.json()
    if payload.get("code", 0) != 0:
        raise RuntimeError(f"Feishu API failed: {payload}")
    return payload


def get_tenant_access_token(app_id: str, app_secret: str) -> str:
    payload = feishu_request(
        "POST",
        "/auth/v3/tenant_access_token/internal",
        json_body={"app_id": app_id, "app_secret": app_secret},
    )
    token = payload.get("tenant_access_token")
    if not token:
        raise RuntimeError(f"Feishu token response missing tenant_access_token: {payload}")
    return str(token)


def list_fields(token: str, app_token: str, table_id: str) -> dict[str, dict[str, Any]]:
    fields: dict[str, dict[str, Any]] = {}
    page_token = ""

    while True:
        params: dict[str, Any] = {"page_size": 100}
        if page_token:
            params["page_token"] = page_token
        payload = feishu_request(
            "GET",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/fields",
            token=token,
            params=params,
        )
        data = payload.get("data") or {}
        for item in data.get("items") or []:
            fields[str(item.get("field_name"))] = item
        if not data.get("has_more"):
            return fields
        page_token = str(data.get("page_token") or "")


def create_field(token: str, app_token: str, table_id: str, field_name: str, field_type: int) -> None:
    body: dict[str, Any] = {"field_name": field_name, "type": field_type}
    if field_type == FIELD_TYPE_NUMBER:
        body["property"] = {"formatter": "0.00"}
    elif field_type == FIELD_TYPE_DATE:
        body["property"] = {"date_formatter": "yyyy/MM/dd HH:mm:ss"}

    feishu_request(
        "POST",
        f"/bitable/v1/apps/{app_token}/tables/{table_id}/fields",
        token=token,
        json_body=body,
    )


def ensure_fields(
    token: str,
    app_token: str,
    table_id: str,
    headers: list[str],
    inferred_types: dict[str, int],
    create_missing_fields: bool,
) -> dict[str, dict[str, Any]]:
    fields = list_fields(token, app_token, table_id)
    missing = [header for header in headers if header not in fields]
    if missing and not create_missing_fields:
        raise RuntimeError("Bitable is missing fields: " + ", ".join(missing))

    for header in missing:
        create_field(token, app_token, table_id, header, inferred_types.get(header, FIELD_TYPE_TEXT))
        print(f"created field: {header}")

    return list_fields(token, app_token, table_id)


def convert_value(value: Any, field_type: int) -> Any:
    if value is None:
        return None
    if value == "":
        return ""
    if field_type == FIELD_TYPE_NUMBER:
        return value if is_number(value) else float(str(value).replace(",", ""))
    if field_type == FIELD_TYPE_DATE:
        parsed = parse_datetime_value(value)
        return datetime_to_millis(parsed) if parsed else None
    return str(value)


def build_bitable_records(
    rows: list[dict[str, Any]],
    fields: dict[str, dict[str, Any]],
) -> list[dict[str, dict[str, Any]]]:
    result: list[dict[str, dict[str, Any]]] = []
    for row in rows:
        bitable_fields: dict[str, Any] = {}
        for header, value in row.items():
            field = fields.get(header) or {}
            field_type = int(field.get("type") or FIELD_TYPE_TEXT)
            converted = convert_value(value, field_type)
            if converted is not None:
                bitable_fields[header] = converted
        result.append({"fields": bitable_fields})
    return result


def chunks(values: list[Any], size: int) -> list[list[Any]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def list_record_ids(token: str, app_token: str, table_id: str) -> list[str]:
    record_ids: list[str] = []
    page_token = ""

    while True:
        params: dict[str, Any] = {"page_size": 500}
        if page_token:
            params["page_token"] = page_token
        payload = feishu_request(
            "GET",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records",
            token=token,
            params=params,
        )
        data = payload.get("data") or {}
        for item in data.get("items") or []:
            record_id = item.get("record_id")
            if record_id:
                record_ids.append(str(record_id))
        if not data.get("has_more"):
            return record_ids
        page_token = str(data.get("page_token") or "")


def clear_records(token: str, app_token: str, table_id: str, batch_size: int, request_interval: float) -> int:
    record_ids = list_record_ids(token, app_token, table_id)
    for batch in chunks(record_ids, min(batch_size, 500)):
        feishu_request(
            "POST",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_delete",
            token=token,
            json_body={"records": batch},
        )
        if request_interval > 0:
            time.sleep(request_interval)
    return len(record_ids)


def upload_records(
    token: str,
    app_token: str,
    table_id: str,
    records: list[dict[str, dict[str, Any]]],
    batch_size: int,
    request_interval: float,
) -> int:
    uploaded = 0
    for batch in chunks(records, batch_size):
        feishu_request(
            "POST",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_create",
            token=token,
            json_body={"records": batch},
        )
        uploaded += len(batch)
        print(f"uploaded records: {uploaded}/{len(records)}")
        if request_interval > 0:
            time.sleep(request_interval)
    return uploaded


def upload_excel_sheet(config: UploadConfig) -> int:
    headers, rows, inferred_types = read_sheet(config.input_path, config.sheet_name)
    print(f"input file: {config.input_path}")
    print(f"sheet: {config.sheet_name}")
    print(f"columns: {len(headers)}")
    print(f"records: {len(rows)}")

    if config.dry_run:
        print("dry run only; no Feishu API calls were made")
        return 0

    token = get_tenant_access_token(config.app_id, config.app_secret)
    fields = ensure_fields(
        token,
        config.app_token,
        config.table_id,
        headers,
        inferred_types,
        config.create_missing_fields,
    )
    records = build_bitable_records(rows, fields)

    if config.clear_table:
        deleted = clear_records(token, config.app_token, config.table_id, config.batch_size, config.request_interval)
        print(f"cleared records: {deleted}")

    return upload_records(token, config.app_token, config.table_id, records, config.batch_size, config.request_interval)


def config_from_args(args: argparse.Namespace) -> UploadConfig:
    load_dotenv(Path(args.env_file).resolve())

    app_token = args.app_token or os.environ.get("FEISHU_APP_TOKEN", "").strip()
    table_id = args.table_id or os.environ.get("FEISHU_TABLE_ID", "").strip()

    bitable_url = args.bitable_url or os.environ.get("FEISHU_BITABLE_URL", "").strip()
    if bitable_url and (not app_token or not table_id):
        parsed_app_token, parsed_table_id = parse_bitable_url(bitable_url)
        app_token = app_token or parsed_app_token
        table_id = table_id or parsed_table_id or ""

    if args.dry_run:
        app_token = app_token or "dry_run_app_token"
        table_id = table_id or "dry_run_table_id"
    elif not app_token:
        raise RuntimeError("Missing app_token. Set FEISHU_BITABLE_URL or FEISHU_APP_TOKEN.")
    elif not table_id:
        raise RuntimeError("Missing table_id. Add ?table=<table_id> to FEISHU_BITABLE_URL or set FEISHU_TABLE_ID.")

    return UploadConfig(
        app_id=args.app_id or (os.environ.get("FEISHU_APP_ID", "").strip() if args.dry_run else require_env("FEISHU_APP_ID")),
        app_secret=args.app_secret
        or (os.environ.get("FEISHU_APP_SECRET", "").strip() if args.dry_run else require_env("FEISHU_APP_SECRET")),
        app_token=app_token,
        table_id=table_id,
        input_path=Path(args.input).resolve(),
        sheet_name=args.sheet,
        batch_size=args.batch_size,
        create_missing_fields=not args.no_create_missing_fields,
        clear_table=args.clear_table,
        dry_run=args.dry_run,
        request_interval=args.request_interval,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Upload a cleaned KPI Excel sheet to Feishu Bitable.")
    parser.add_argument("input", help="Cleaned Excel file path")
    parser.add_argument("--sheet", default=os.environ.get("FEISHU_SHEET_NAME", DEFAULT_SHEET_NAME))
    parser.add_argument("--bitable-url", default="")
    parser.add_argument("--app-token", default="")
    parser.add_argument("--table-id", default="")
    parser.add_argument("--app-id", default="")
    parser.add_argument("--app-secret", default="")
    parser.add_argument("--env-file", default=".env")
    parser.add_argument("--batch-size", type=int, default=500)
    parser.add_argument("--request-interval", type=float, default=0.2)
    parser.add_argument("--clear-table", action="store_true", help="Delete all existing records before upload")
    parser.add_argument("--no-create-missing-fields", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    uploaded = upload_excel_sheet(config_from_args(args))
    if not args.dry_run:
        print(f"done, uploaded records: {uploaded}")


if __name__ == "__main__":
    main()
