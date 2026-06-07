"""KPI #2: 拣货人绩效统计

基于出库单 Excel 数据，按拣货人维度统计绩效指标：
  - 波次数、订单数、SKU 数、总数量
  - 最早/最晚拣货时间、工作时长
  - 拣货效率（件/小时、单/小时）

用法:
  python clean_picker_performance_kpi.py delivery_20260608.xlsx --kpi-date 2026-06-08 --fetch-picker-names --storage-state storage_state.json
"""
from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from omp_api import resolve_picker_names

# ── Sheet 名称 ──────────────────────────────────────────────
MAIN_SHEET = "出库单"
DETAIL_SHEET = "拣货库位明细"
SUMMARY_SHEET = "拣货人绩效汇总"
DAILY_SHEET = "拣货人日报"
RANKING_SHEET = "拣货人排名"

# ── 可能的列名（中英文双语）─────────────────────────────────
WAVE_NO_CANDIDATES = [
    "Wave No./关联波次号",
    "Wave No.",
    "关联波次号",
    "waveNo",
]

PICK_TIME_CANDIDATES = [
    "Pick time/拣货时间",
    "Pick time",
    "拣货时间",
    "pickTime",
]

TOTAL_QTY_CANDIDATES = [
    "Total Qty of SKU/总数量",
    "Total Qty of SKU",
    "总数量",
    "totalQty",
]

ORDER_NO_CANDIDATES = [
    "Delivery No./出库单号",
    "Delivery No.",
    "出库单号",
    "deliveryNo",
    "Order No./订单号",
    "Order No.",
    "订单号",
]

SKU_CODE_CANDIDATES = [
    "SKU/产品编码",
    "SKU",
    "产品编码",
    "skuCode",
    "Seller SKU/商家SKU",
    "Seller SKU",
    "商家SKU",
]

CREATE_TIME_CANDIDATES = [
    "Create time/创建时间",
    "Create time",
    "创建时间",
    "createTime",
    "Order time/下单时间",
    "Order time",
    "下单时间",
]

SHIP_TIME_CANDIDATES = [
    "Ship time/出库时间",
    "Ship time",
    "出库时间",
    "shipTime",
    "Delivery time/发货时间",
    "Delivery time",
    "发货时间",
]

# ── 绩效汇总表头 ────────────────────────────────────────────
PERFORMANCE_HEADERS = [
    "日期",
    "周",
    "月",
    "拣货人",
    "波次数",
    "订单数",
    "SKU数",
    "总数量",
    "最早拣货时间",
    "最晚拣货时间",
    "工作时长(小时)",
    "拣货效率(件/小时)",
    "拣货效率(单/小时)",
]

DAILY_HEADERS = [
    "日期",
    "周",
    "拣货人",
    "波次数",
    "订单数",
    "SKU数",
    "总数量",
    "最早拣货时间",
    "最晚拣货时间",
    "工作时长(小时)",
    "拣货效率(件/小时)",
]

RANKING_HEADERS = [
    "排名",
    "拣货人",
    "日期",
    "总数量",
    "订单数",
    "波次数",
    "拣货效率(件/小时)",
    "工作时长(小时)",
]

# ── 样式 ────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(size=10)
NUMBER_FONT = Font(size=10)
GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
SILVER_FILL = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
BRONZE_FILL = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")


# ── 辅助函数 ────────────────────────────────────────────────
def compact_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def find_header(headers: list[Any], candidates: list[str]) -> int:
    """从候选列名中匹配 Excel 列索引"""
    normalized_candidates = [compact_header(c) for c in candidates]
    for index, header in enumerate(headers):
        if compact_header(header) in normalized_candidates:
            return index
    raise RuntimeError(f"Excel missing required column. Tried: {candidates}")


def find_header_optional(headers: list[Any], candidates: list[str]) -> int | None:
    """可选的列匹配，找不到返回 None"""
    try:
        return find_header(headers, candidates)
    except RuntimeError:
        return None


def parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def format_date(value: date) -> str:
    return value.strftime("%Y/%m/%d")


def format_datetime(value: datetime | None) -> str:
    return "" if value is None else value.strftime("%Y/%m/%d %H:%M:%S")


def round2(value: float) -> float:
    return round(value, 2)


def picker_search_window(kpi_day: date) -> tuple[datetime, datetime]:
    """拣货人查询的时间窗口（同 KPI 窗口）"""
    start_at = datetime.combine(kpi_day - timedelta(days=1), time(22, 30, 0))
    if kpi_day.weekday() == 4:
        end_at = datetime.combine(kpi_day, time(23, 59, 59))
    else:
        end_at = datetime.combine(kpi_day, time(22, 30, 0))
    return start_at, end_at


def pick_time_exclusion_window(kpi_day: date) -> tuple[datetime, datetime]:
    """拣货时间排除窗口（前一天 22:30 ~ 当天 10:30 的数据不参与绩效统计）"""
    start_at = datetime.combine(kpi_day - timedelta(days=1), time(22, 30, 0))
    end_at = datetime.combine(kpi_day, time(10, 30, 0))
    return start_at, end_at


# ── 样式设置 ────────────────────────────────────────────────
def style_sheet(ws: Any, header_fill=None, header_font=None) -> None:
    """统一设置 sheet 样式"""
    if header_fill is None:
        header_fill = HEADER_FILL
    if header_font is None:
        header_font = HEADER_FONT
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.tabColor = "4472C4"


def auto_width(ws: Any, max_width: int = 36) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, max_width))
        ws.column_dimensions[column_letter].width = width


# ── 数据提取 ────────────────────────────────────────────────
def extract_rows_with_picker(
    workbook: Any,
    kpi_day: date,
    picker_names: dict[str, str],
) -> list[dict[str, Any]]:
    """从出库单 sheet 提取每行数据，附加拣货人和过滤逻辑"""
    ws = workbook[MAIN_SHEET]
    headers = [cell.value for cell in ws[1]]

    wave_idx = find_header(headers, WAVE_NO_CANDIDATES)
    pick_time_idx = find_header(headers, PICK_TIME_CANDIDATES)
    qty_idx = find_header(headers, TOTAL_QTY_CANDIDATES)
    order_no_idx = find_header_optional(headers, ORDER_NO_CANDIDATES)
    sku_idx = find_header_optional(headers, SKU_CODE_CANDIDATES)
    create_time_idx = find_header_optional(headers, CREATE_TIME_CANDIDATES)
    ship_time_idx = find_header_optional(headers, SHIP_TIME_CANDIDATES)

    exclude_start, exclude_end = pick_time_exclusion_window(kpi_day)

    rows = []
    kept = 0
    removed = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        wave_no = str(row[wave_idx] or "").strip()
        pick_time = parse_datetime(row[pick_time_idx])

        # 排除窗口内的数据不参与统计
        should_remove = pick_time is not None and exclude_start <= pick_time <= exclude_end
        if should_remove:
            removed += 1
            continue
        kept += 1

        picker = picker_names.get(wave_no, "")

        qty = parse_number(row[qty_idx])

        rows.append({
            "wave_no": wave_no,
            "picker": picker,
            "pick_time": pick_time,
            "qty": qty or 0,
            "order_no": str(row[order_no_idx] or "").strip() if order_no_idx is not None else "",
            "sku_code": str(row[sku_idx] or "").strip() if sku_idx is not None else "",
            "create_time": parse_datetime(row[create_time_idx]) if create_time_idx is not None else None,
            "ship_time": parse_datetime(row[ship_time_idx]) if ship_time_idx is not None else None,
        })

    print(f"Pick time exclusion window: {exclude_start} ~ {exclude_end}")
    print(f"Kept rows: {kept}, removed rows: {removed}")
    return rows


# ── 绩效汇总 ────────────────────────────────────────────────
def aggregate_picker_stats(
    rows: list[dict[str, Any]],
    kpi_day: date,
) -> list[dict[str, Any]]:
    """按拣货人聚合统计"""
    # 按 (拣货人) 聚合整体数据
    picker_data: dict[str, dict[str, Any]] = {}

    for row in rows:
        picker = row["picker"]
        if not picker:
            continue

        if picker not in picker_data:
            picker_data[picker] = {
                "waves": set(),
                "orders": set(),
                "skus": set(),
                "total_qty": 0.0,
                "min_pick": None,
                "max_pick": None,
            }

        stats = picker_data[picker]
        if row["wave_no"]:
            stats["waves"].add(row["wave_no"])
        if row["order_no"]:
            stats["orders"].add(row["order_no"])
        if row["sku_code"]:
            stats["skus"].add(row["sku_code"])
        stats["total_qty"] += row["qty"]

        pt = row["pick_time"]
        if pt is not None:
            if stats["min_pick"] is None or pt < stats["min_pick"]:
                stats["min_pick"] = pt
            if stats["max_pick"] is None or pt > stats["max_pick"]:
                stats["max_pick"] = pt

    iso_year, iso_week, _ = kpi_day.isocalendar()
    report_date = format_date(kpi_day)
    report_week = f"{iso_year}-W{iso_week:02d}"
    report_month = kpi_day.replace(day=1).strftime("%Y/%m/%d")

    results = []
    for picker, stats in picker_data.items():
        wave_count = len(stats["waves"])
        order_count = len(stats["orders"])
        sku_count = len(stats["skus"])
        total_qty = int(stats["total_qty"]) if stats["total_qty"].is_integer() else round2(stats["total_qty"])

        work_hours = 0.0
        if stats["min_pick"] and stats["max_pick"]:
            diff = (stats["max_pick"] - stats["min_pick"]).total_seconds() / 3600.0
            work_hours = max(round2(diff), 0.01)

        efficiency_per_hour = round2(total_qty / work_hours) if work_hours > 0 else 0.0
        orders_per_hour = round2(order_count / work_hours) if work_hours > 0 else 0.0

        results.append({
            "date": report_date,
            "week": report_week,
            "month": report_month,
            "picker": picker,
            "wave_count": wave_count,
            "order_count": order_count,
            "sku_count": sku_count,
            "total_qty": total_qty,
            "min_pick": format_datetime(stats["min_pick"]),
            "max_pick": format_datetime(stats["max_pick"]),
            "work_hours": work_hours,
            "efficiency_per_hour": efficiency_per_hour,
            "orders_per_hour": orders_per_hour,
            # 保留原始值用于排名
            "_min_pick_raw": stats["min_pick"],
            "_max_pick_raw": stats["max_pick"],
            "_total_qty_raw": total_qty,
            "_efficiency_raw": efficiency_per_hour,
            "_work_hours_raw": work_hours,
            "_order_count_raw": order_count,
            "_wave_count_raw": wave_count,
        })

    # 按总数量降序
    results.sort(key=lambda x: x["_total_qty_raw"], reverse=True)
    return results


# ── 每日统计 ────────────────────────────────────────────────
def aggregate_daily_stats(
    rows: list[dict[str, Any]],
    kpi_day: date,
) -> list[dict[str, Any]]:
    """按 (日期, 拣货人) 聚合每日统计"""
    # 简化：当前只处理单个 kpi_day，按拣货人聚合
    # 与汇总类似，但为扩展多日统计预留结构
    return aggregate_picker_stats(rows, kpi_day)


# ── 生成 Excel sheet ───────────────────────────────────────
def build_summary_sheet(workbook: Any, stats: list[dict[str, Any]]) -> None:
    """生成拣货人绩效汇总 sheet"""
    if SUMMARY_SHEET in workbook.sheetnames:
        del workbook[SUMMARY_SHEET]
    ws = workbook.create_sheet(SUMMARY_SHEET, 0)
    ws.append(PERFORMANCE_HEADERS)

    for s in stats:
        ws.append([
            s["date"], s["week"], s["month"], s["picker"],
            s["wave_count"], s["order_count"], s["sku_count"], s["total_qty"],
            s["min_pick"], s["max_pick"],
            s["work_hours"], s["efficiency_per_hour"], s["orders_per_hour"],
        ])
    style_sheet(ws)
    auto_width(ws)


def build_ranking_sheet(workbook: Any, stats: list[dict[str, Any]], kpi_day: date) -> None:
    """生成拣货人排名 sheet"""
    if RANKING_SHEET in workbook.sheetnames:
        del workbook[RANKING_SHEET]
    ws = workbook.create_sheet(RANKING_SHEET)
    ws.append(RANKING_HEADERS)

    # 按效率排名
    ranked = sorted(stats, key=lambda x: x["_efficiency_raw"], reverse=True)
    for rank, s in enumerate(ranked, 1):
        row_data = [
            rank, s["picker"], s["date"],
            s["_total_qty_raw"], s["_order_count_raw"], s["_wave_count_raw"],
            s["_efficiency_raw"], s["_work_hours_raw"],
        ]
        ws.append(row_data)

    # 高亮前三名
    for rank, row_idx in enumerate(range(2, len(ranked) + 2), 1):
        fill = None
        if rank == 1:
            fill = GOLD_FILL
        elif rank == 2:
            fill = SILVER_FILL
        elif rank == 3:
            fill = BRONZE_FILL
        if fill:
            for cell in ws[row_idx]:
                cell.fill = fill

    style_sheet(ws)
    auto_width(ws)


def build_daily_sheet(workbook: Any, stats: list[dict[str, Any]]) -> None:
    """生成拣货人日报 sheet"""
    if DAILY_SHEET in workbook.sheetnames:
        del workbook[DAILY_SHEET]
    ws = workbook.create_sheet(DAILY_SHEET)
    ws.append(DAILY_HEADERS)

    for s in stats:
        ws.append([
            s["date"], s["week"], s["picker"],
            s["wave_count"], s["order_count"], s["sku_count"], s["total_qty"],
            s["min_pick"], s["max_pick"],
            s["work_hours"], s["efficiency_per_hour"],
        ])
    style_sheet(ws)
    auto_width(ws)


# ── 主流程 ──────────────────────────────────────────────────
def clean_picker_performance(
    input_path: Path,
    output_path: Path,
    kpi_day: date,
    args: argparse.Namespace,
) -> None:
    workbook = load_workbook(input_path)

    # 确保出库单 sheet 存在
    if MAIN_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"Excel cannot find main sheet: {MAIN_SHEET}")

    # 解析拣货人
    picker_names: dict[str, str] = {}
    if args.fetch_picker_names:
        # 先收集所有波次号
        ws = workbook[MAIN_SHEET]
        headers = [cell.value for cell in ws[1]]
        wave_idx = find_header(headers, WAVE_NO_CANDIDATES)
        wave_nos_set: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            wave_no = str(row[wave_idx] or "").strip()
            if wave_no:
                wave_nos_set.add(wave_no)
        wave_nos = list(wave_nos_set)

        default_start, default_end = picker_search_window(kpi_day)
        start_time = args.picker_start_time or default_start.strftime("%Y-%m-%d %H:%M:%S")
        end_time = args.picker_end_time or default_end.strftime("%Y-%m-%d %H:%M:%S")
        picker_names = resolve_picker_names(
            wave_nos=wave_nos,
            storage_state=args.storage_state,
            start_time=start_time,
            end_time=end_time,
            page_size=args.picker_page_size,
            sample_limit=args.picker_sample_limit,
            wave_limit=args.picker_wave_limit,
            sleep_seconds=args.picker_sleep,
        )

    # 提取数据（附带拣货人）
    rows = extract_rows_with_picker(workbook, kpi_day, picker_names)

    # 聚合统计
    stats = aggregate_picker_stats(rows, kpi_day)

    # 生成 sheet
    build_summary_sheet(workbook, stats)
    build_ranking_sheet(workbook, stats, kpi_day)
    build_daily_sheet(workbook, stats)

    # 保留出库单 sheet，删除拣货库位明细
    if DETAIL_SHEET in workbook.sheetnames:
        del workbook[DETAIL_SHEET]

    # 删除其他无关 sheet
    for sheet_name in list(workbook.sheetnames):
        if sheet_name not in (MAIN_SHEET, SUMMARY_SHEET, RANKING_SHEET, DAILY_SHEET):
            del workbook[sheet_name]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    # 打印统计
    print(f"\n{'='*60}")
    print(f"KPI date: {kpi_day:%Y-%m-%d}")
    print(f"Total pickers: {len(stats)}")
    print(f"Total rows processed: {len(rows)}")
    if args.fetch_picker_names:
        found = sum(1 for v in picker_names.values() if v)
        print(f"Picker names resolved: {found}/{len(picker_names)}")
    print(f"\nTop 5 pickers by quantity:")
    for i, s in enumerate(stats[:5], 1):
        print(f"  {i}. {s['picker']}: {s['total_qty']} items, {s['order_count']} orders, "
              f"{s['efficiency_per_hour']} items/hr, {s['work_hours']}h")
    print(f"\nOutput: {output_path.resolve()}")
    print(f"{'='*60}")


def main() -> None:
    parser = argparse.ArgumentParser(description="拣货人绩效 KPI 统计")
    parser.add_argument("input", help="Raw delivery export Excel path")
    parser.add_argument("--kpi-date", required=True, help="KPI date, YYYY-MM-DD")
    parser.add_argument("--output", default="", help="Output Excel path")
    parser.add_argument("--fetch-picker-names", action="store_true", help="Fetch picker names from OMP API")
    parser.add_argument("--storage-state", default="", help="Playwright storage_state JSON path")
    parser.add_argument("--picker-start-time", default="", help="Override wave search start time")
    parser.add_argument("--picker-end-time", default="", help="Override wave search end time")
    parser.add_argument("--picker-page-size", type=int, default=20)
    parser.add_argument("--picker-sample-limit", type=int, default=3)
    parser.add_argument("--picker-wave-limit", type=int, default=0, help="0 means all")
    parser.add_argument("--picker-sleep", type=float, default=0.0)
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if args.output:
        output_path = Path(args.output).resolve()
    else:
        output_dir = input_path.parent / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{input_path.stem}_picker_performance.xlsx"

    clean_picker_performance(input_path, output_path, _parse_kpi_date(args.kpi_date), args)


def _parse_kpi_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


if __name__ == "__main__":
    main()
