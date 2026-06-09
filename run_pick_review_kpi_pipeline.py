from __future__ import annotations

import argparse
import sys
from argparse import Namespace
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from clean_delivery_kpi import parse_kpi_date
from clean_outbound_kpi import clean_outbound_workbook
from clean_review_kpi import clean_review_workbook, create_time_window
from export_delivery_once import (
    build_payload,
    create_export_task,
    download_when_ready,
    session_from_args,
    wh_code_from_storage_state,
)


API_PROJECT_DIR = Path(r"C:\Users\Administrator\-API-")

PICK_SUMMARY_SHEET = "拣货人绩效汇总"
REVIEW_SUMMARY_SHEET = "复核人员汇总"
OUTBOUND_SUMMARY_SHEET = "出库人员汇总"
UNIFIED_PERSON_SHEET = "人员KPI汇总"

UNIFIED_HEADERS = [
    "日期",
    "周",
    "月",
    "环节",
    "名字",
    "总数量",
    "订单数",
    "波次数量",
    "最早时间",
    "最晚时间",
    "工作小时数",
    "工作时间",
    "总数占比",
    "波次占比",
    "件/小时",
    "波次/小时",
    "平均每波总数",
    "平均每波订单数",
    "平均每波工作分钟",
]


def compact_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def index_by_header(sheet: Any) -> dict[str, int]:
    return {
        compact_header(value): index
        for index, value in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        if compact_header(value)
    }


def cell(row: tuple[Any, ...], indexes: dict[str, int], name: str, default: Any = "") -> Any:
    index = indexes.get(name)
    if index is None or index >= len(row):
        return default
    value = row[index]
    return default if value is None else value


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def display_number(value: float) -> int | float:
    return int(value) if float(value).is_integer() else round(value, 2)


def format_hours(hours: float) -> str:
    total_minutes = int(round(hours * 60))
    return f"{total_minutes // 60}小时{total_minutes % 60:02d}分"


def style_header(sheet: Any) -> None:
    for item in sheet[1]:
        item.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def resize_columns(sheet: Any, max_width: int = 32) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 10
        for item in column_cells:
            value = "" if item.value is None else str(item.value)
            width = max(width, min(len(value) + 2, max_width))
        sheet.column_dimensions[column_letter].width = width


def default_output_dir(kpi_day) -> Path:
    return Path.cwd() / "output" / "kpi" / kpi_day.strftime("%Y%m%d") / "pick_review"


def download_delivery_excel(args: argparse.Namespace, output_dir: Path) -> Path:
    kpi_day = parse_kpi_date(args.kpi_date)
    create_start, create_end = create_time_window(kpi_day)
    export_start = args.export_start_time or create_start.strftime("%Y-%m-%d %H:%M:%S")
    export_end = args.export_end_time or create_end.strftime("%Y-%m-%d %H:%M:%S")
    storage_state = str(Path(args.storage_state).resolve())
    wh_code = args.wh_code or wh_code_from_storage_state(storage_state)

    print(f"export window: {export_start} ~ {export_end}", flush=True)
    print(f"warehouse: {wh_code or '-'}, status: {args.status}", flush=True)

    session = session_from_args(Namespace(storage_state=storage_state, wh_code=wh_code))
    payload = build_payload(export_start, export_end, wh_code, args.status)
    task_id = create_export_task(session, payload)
    print(f"export task created: {task_id}", flush=True)

    raw_path = output_dir / "raw" / f"delivery_raw_{kpi_day:%Y%m%d}_{task_id}.xlsx"
    return download_when_ready(session, task_id, raw_path, args.attempts, args.interval)


def run_picker_kpi(args: argparse.Namespace, raw_path: Path, output_path: Path) -> Path:
    if not API_PROJECT_DIR.exists():
        raise FileNotFoundError(f"Reference API project not found: {API_PROJECT_DIR}")
    if str(API_PROJECT_DIR) not in sys.path:
        sys.path.insert(0, str(API_PROJECT_DIR))

    import clean_picker_performance_kpi as picker_module

    for header in ("Outbound Order No/出库单号", "Outbound Order No"):
        if header not in picker_module.ORDER_NO_CANDIDATES:
            picker_module.ORDER_NO_CANDIDATES.insert(0, header)

    picker_args = Namespace(
        fetch_picker_names=True,
        storage_state=str(Path(args.storage_state).resolve()),
        picker_start_time=args.picker_start_time,
        picker_end_time=args.picker_end_time,
        picker_page_size=args.picker_page_size,
        picker_sample_limit=args.picker_sample_limit,
        picker_wave_limit=args.picker_wave_limit,
        picker_sleep=args.picker_sleep,
        picker_max_workers=args.picker_max_workers,
        picker_max_pages=args.picker_max_pages,
        picker_retry_rounds=args.picker_retry_rounds,
        picker_cache=args.picker_cache,
        allow_missing_picker_names=True,
    )
    picker_module.clean_picker_performance(raw_path, output_path, parse_kpi_date(args.kpi_date), picker_args)
    return output_path


def run_review_kpi(args: argparse.Namespace, raw_path: Path, output_path: Path) -> Path:
    review_args = Namespace(
        name=args.name,
        fetch_reviewer_names=True,
        storage_state=str(Path(args.storage_state).resolve()),
        order_start_time=args.order_start_time,
        order_end_time=args.order_end_time,
        reviewer_page_size=args.reviewer_page_size,
        reviewer_sample_limit=args.reviewer_sample_limit,
        reviewer_wave_limit=args.reviewer_wave_limit,
        reviewer_sleep=args.reviewer_sleep,
    )
    clean_review_workbook(raw_path, output_path, parse_kpi_date(args.kpi_date), review_args)
    return output_path


def run_outbound_kpi(args: argparse.Namespace, raw_path: Path, output_path: Path) -> Path:
    outbound_args = Namespace(
        name=args.name,
        fetch_outbound_names=True,
        storage_state=str(Path(args.storage_state).resolve()),
        order_start_time=args.order_start_time,
        order_end_time=args.order_end_time,
        outbound_page_size=args.outbound_page_size,
        outbound_sample_limit=args.outbound_sample_limit,
        outbound_wave_limit=args.outbound_wave_limit,
        outbound_sleep=args.outbound_sleep,
    )
    clean_outbound_workbook(raw_path, output_path, parse_kpi_date(args.kpi_date), outbound_args)
    return output_path


def read_picker_summary_rows(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if PICK_SUMMARY_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"{path} missing sheet: {PICK_SUMMARY_SHEET}")
    sheet = workbook[PICK_SUMMARY_SHEET]
    indexes = index_by_header(sheet)

    raw_rows = [
        row
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in row)
    ]
    total_qty_all = sum(number(cell(row, indexes, "总数量")) for row in raw_rows)
    total_waves_all = sum(number(cell(row, indexes, "波次数")) for row in raw_rows)

    result: list[list[Any]] = []
    for row in raw_rows:
        qty = number(cell(row, indexes, "总数量"))
        orders = number(cell(row, indexes, "订单数"))
        waves = number(cell(row, indexes, "波次数"))
        hours = number(cell(row, indexes, "工作时长(小时)"))

        result.append(
            [
                cell(row, indexes, "日期"),
                cell(row, indexes, "周"),
                cell(row, indexes, "月"),
                "拣货",
                cell(row, indexes, "拣货人", "未匹配"),
                display_number(qty),
                display_number(orders),
                display_number(waves),
                cell(row, indexes, "最早拣货时间"),
                cell(row, indexes, "最晚拣货时间"),
                hours,
                format_hours(hours),
                round(qty / total_qty_all, 6) if total_qty_all else "",
                round(waves / total_waves_all, 6) if total_waves_all else "",
                cell(row, indexes, "拣货效率(件/小时)"),
                cell(row, indexes, "拣货效率(单/小时)"),
                round(qty / waves, 2) if waves else "",
                round(orders / waves, 2) if waves else "",
                round(hours * 60 / waves, 2) if waves else "",
            ]
        )
    return result


def read_review_summary_rows(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if REVIEW_SUMMARY_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"{path} missing sheet: {REVIEW_SUMMARY_SHEET}")
    sheet = workbook[REVIEW_SUMMARY_SHEET]
    indexes = index_by_header(sheet)

    result: list[list[Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        result.append(
            [
                cell(row, indexes, "日期"),
                cell(row, indexes, "周"),
                cell(row, indexes, "月"),
                cell(row, indexes, "环节", "复核-一单一件"),
                cell(row, indexes, "名字", "未匹配"),
                cell(row, indexes, "总数量"),
                cell(row, indexes, "订单数"),
                cell(row, indexes, "波次数量"),
                cell(row, indexes, "最早复核时间"),
                cell(row, indexes, "最晚复核时间"),
                cell(row, indexes, "工作小时数"),
                cell(row, indexes, "工作时间"),
                cell(row, indexes, "总数占比"),
                cell(row, indexes, "波次占比"),
                cell(row, indexes, "件/小时"),
                cell(row, indexes, "波次/小时"),
                cell(row, indexes, "平均每波总数"),
                cell(row, indexes, "平均每波订单数"),
                cell(row, indexes, "平均每波工作分钟"),
            ]
        )
    return result


def read_outbound_summary_rows(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if OUTBOUND_SUMMARY_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"{path} missing sheet: {OUTBOUND_SUMMARY_SHEET}")
    sheet = workbook[OUTBOUND_SUMMARY_SHEET]
    indexes = index_by_header(sheet)

    result: list[list[Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        result.append(
            [
                cell(row, indexes, "日期"),
                cell(row, indexes, "周"),
                cell(row, indexes, "月"),
                cell(row, indexes, "环节", "出库-一单多件"),
                cell(row, indexes, "名字", "未匹配"),
                cell(row, indexes, "总数量"),
                cell(row, indexes, "订单数"),
                cell(row, indexes, "波次数量"),
                cell(row, indexes, "最早出库时间"),
                cell(row, indexes, "最晚出库时间"),
                cell(row, indexes, "工作小时数"),
                cell(row, indexes, "工作时间"),
                cell(row, indexes, "总数占比"),
                cell(row, indexes, "波次占比"),
                cell(row, indexes, "件/小时"),
                cell(row, indexes, "波次/小时"),
                cell(row, indexes, "平均每波总数"),
                cell(row, indexes, "平均每波订单数"),
                cell(row, indexes, "平均每波工作分钟"),
            ]
        )
    return result


def build_unified_summary(picker_path: Path, review_path: Path, outbound_path: Path, output_path: Path) -> int:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = UNIFIED_PERSON_SHEET
    sheet.append(UNIFIED_HEADERS)

    rows = read_picker_summary_rows(picker_path) + read_review_summary_rows(review_path) + read_outbound_summary_rows(outbound_path)
    for row in rows:
        sheet.append(row)

    style_header(sheet)
    resize_columns(sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(rows)


def run_pipeline(args: argparse.Namespace) -> None:
    kpi_day = parse_kpi_date(args.kpi_date)
    output_dir = Path(args.output_dir).resolve() if args.output_dir else default_output_dir(kpi_day).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"KPI date: {args.kpi_date}", flush=True)
    if args.input:
        raw_path = Path(args.input).resolve()
        print(f"download skipped, using input: {raw_path}", flush=True)
    else:
        raw_path = download_delivery_excel(args, output_dir)
        print(f"downloaded raw file: {raw_path}", flush=True)

    picker_path = output_dir / "pick" / f"pick_kpi_{kpi_day:%Y%m%d}.xlsx"
    review_path = output_dir / "review" / f"review_kpi_{kpi_day:%Y%m%d}.xlsx"
    outbound_path = output_dir / "outbound" / f"outbound_kpi_{kpi_day:%Y%m%d}.xlsx"
    summary_path = output_dir / "summary" / f"person_kpi_summary_{kpi_day:%Y%m%d}.xlsx"

    print("\n=== 1/4 拣货 KPI ===", flush=True)
    run_picker_kpi(args, raw_path, picker_path)

    print("\n=== 2/4 复核 KPI ===", flush=True)
    run_review_kpi(args, raw_path, review_path)

    print("\n=== 3/4 出库 KPI ===", flush=True)
    run_outbound_kpi(args, raw_path, outbound_path)

    print("\n=== 4/4 人员KPI汇总 ===", flush=True)
    rows = build_unified_summary(picker_path, review_path, outbound_path, summary_path)

    print("\nDONE", flush=True)
    print(f"raw file: {raw_path}", flush=True)
    print(f"pick file: {picker_path}", flush=True)
    print(f"review file: {review_path}", flush=True)
    print(f"outbound file: {outbound_path}", flush=True)
    print(f"summary rows: {rows}", flush=True)
    print(f"summary file: {summary_path}", flush=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Download and run pick + review + outbound KPI, then build unified person summary.")
    parser.add_argument("--kpi-date", required=True)
    parser.add_argument("--input", default="", help="Use existing delivery Excel instead of downloading")
    parser.add_argument("--storage-state", default="storage_state.json")
    parser.add_argument("--output-dir", default="")
    parser.add_argument("--name", default="未匹配")

    parser.add_argument("--export-start-time", default="")
    parser.add_argument("--export-end-time", default="")
    parser.add_argument("--wh-code", default="")
    parser.add_argument("--status", default="100")
    parser.add_argument("--attempts", type=int, default=30)
    parser.add_argument("--interval", type=int, default=10)

    parser.add_argument("--picker-start-time", default="")
    parser.add_argument("--picker-end-time", default="")
    parser.add_argument("--picker-page-size", type=int, default=50)
    parser.add_argument("--picker-sample-limit", type=int, default=2)
    parser.add_argument("--picker-wave-limit", type=int, default=0)
    parser.add_argument("--picker-sleep", type=float, default=0.0)
    parser.add_argument("--picker-max-workers", type=int, default=4)
    parser.add_argument("--picker-max-pages", type=int, default=50)
    parser.add_argument("--picker-retry-rounds", type=int, default=5)
    parser.add_argument("--picker-cache", default="output/picker_names_cache.json")

    parser.add_argument("--order-start-time", default="")
    parser.add_argument("--order-end-time", default="")
    parser.add_argument("--reviewer-page-size", type=int, default=100)
    parser.add_argument("--reviewer-sample-limit", type=int, default=3)
    parser.add_argument("--reviewer-wave-limit", type=int, default=0)
    parser.add_argument("--reviewer-sleep", type=float, default=0.0)

    parser.add_argument("--outbound-page-size", type=int, default=100)
    parser.add_argument("--outbound-sample-limit", type=int, default=3)
    parser.add_argument("--outbound-wave-limit", type=int, default=0)
    parser.add_argument("--outbound-sleep", type=float, default=0.0)
    args = parser.parse_args()

    run_pipeline(args)


if __name__ == "__main__":
    main()
