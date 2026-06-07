"""KPI #3: 出库时效分析

基于出库单 Excel 数据，分析订单各阶段时效：
  - 下单→拣货 耗时
  - 拣货→发货 耗时
  - 下单→发货 总耗时
  - SLA 达成率（12h / 24h / 48h 出库率）
  - 按波次号、客户、仓库维度的时效分布

用法:
  python clean_delivery_leadtime_kpi.py delivery_20260608.xlsx --kpi-date 2026-06-08
"""
from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Sheet 名称 ──────────────────────────────────────────────
MAIN_SHEET = "出库单"
LEADTIME_SUMMARY = "时效汇总"
LEADTIME_BY_WAVE = "按波次时效"
LEADTIME_DISTRIBUTION = "时效分布"
SLA_SHEET = "SLA达成率"

# ── 候选列名 ────────────────────────────────────────────────
WAVE_NO_CANDIDATES = [
    "Wave No./关联波次号", "Wave No.", "关联波次号", "waveNo",
]

PICK_TIME_CANDIDATES = [
    "Pick time/拣货时间", "Pick time", "拣货时间", "pickTime",
]

CREATE_TIME_CANDIDATES = [
    "Create time/创建时间", "Create time", "创建时间", "createTime",
    "Order time/下单时间", "Order time", "下单时间", "orderTime",
]

SHIP_TIME_CANDIDATES = [
    "Ship time/出库时间", "Ship time", "出库时间", "shipTime",
    "Delivery time/发货时间", "Delivery time", "发货时间", "deliveryTime",
]

ORDER_NO_CANDIDATES = [
    "Delivery No./出库单号", "Delivery No.", "出库单号", "deliveryNo",
]

CUSTOMER_CODE_CANDIDATES = [
    "Customer Code/客户编码", "Customer Code", "客户编码", "customerCode",
]

WH_CODE_CANDIDATES = [
    "WH Code/仓库编码", "WH Code", "仓库编码", "whCode",
]

STATUS_CANDIDATES = [
    "Status/状态", "Status", "状态", "status",
]

# ── 表头 ────────────────────────────────────────────────────
SUMMARY_HEADERS = [
    "日期", "周", "月",
    "总订单数", "有创建时间订单数", "有发货时间订单数",
    "平均下单→拣货(小时)", "中位数下单→拣货(小时)",
    "平均拣货→发货(小时)", "中位数拣货→发货(小时)",
    "平均下单→发货(小时)", "中位数下单→发货(小时)",
    "最快下单→发货(小时)", "最慢下单→发货(小时)",
    "12h出库率", "24h出库率", "48h出库率",
]

WAVE_HEADERS = [
    "波次号", "订单数",
    "平均下单→拣货(h)", "平均拣货→发货(h)", "平均下单→发货(h)",
    "最快(h)", "最慢(h)",
]

DISTRIBUTION_HEADERS = [
    "时效区间", "订单数", "占比",
]

SLA_HEADERS = [
    "SLA标准", "达标订单数", "总订单数", "达标率",
]

# ── 样式 ────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
RED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")


# ── 辅助函数 ────────────────────────────────────────────────
def compact_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def find_header(headers: list[Any], candidates: list[str]) -> int:
    normalized_candidates = [compact_header(c) for c in candidates]
    for index, header in enumerate(headers):
        if compact_header(header) in normalized_candidates:
            return index
    raise RuntimeError(f"Excel missing required column. Tried: {candidates}")


def find_header_optional(headers: list[Any], candidates: list[str]) -> int | None:
    try:
        return find_header(headers, candidates)
    except RuntimeError:
        return None


def parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def format_date(value: date) -> str:
    return value.strftime("%Y/%m/%d")


def format_datetime(value: datetime | None) -> str:
    return "" if value is None else value.strftime("%Y/%m/%d %H:%M:%S")


def round2(value: float) -> float:
    return round(value, 2)


def hours_between(start_dt: datetime, end_dt: datetime) -> float:
    """计算两个时间之间的小时数"""
    return round2((end_dt - start_dt).total_seconds() / 3600.0)


def median(values: list[float]) -> float:
    """计算中位数"""
    if not values:
        return 0.0
    sorted_vals = sorted(values)
    n = len(sorted_vals)
    if n % 2 == 0:
        return round2((sorted_vals[n // 2 - 1] + sorted_vals[n // 2]) / 2)
    else:
        return round2(sorted_vals[n // 2])


# ── 样式设置 ────────────────────────────────────────────────
def style_sheet(ws: Any) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_width(ws: Any, max_width: int = 36) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, max_width))
        ws.column_dimensions[column_letter].width = width


# ── 数据提取 ────────────────────────────────────────────────
def extract_order_timeline(workbook: Any) -> list[dict[str, Any]]:
    """提取每条订单的时间线数据"""
    ws = workbook[MAIN_SHEET]
    headers = [cell.value for cell in ws[1]]

    wave_idx = find_header(headers, WAVE_NO_CANDIDATES)
    pick_time_idx = find_header(headers, PICK_TIME_CANDIDATES)
    create_time_idx = find_header_optional(headers, CREATE_TIME_CANDIDATES)
    ship_time_idx = find_header_optional(headers, SHIP_TIME_CANDIDATES)
    order_no_idx = find_header_optional(headers, ORDER_NO_CANDIDATES)
    customer_idx = find_header_optional(headers, CUSTOMER_CODE_CANDIDATES)
    wh_idx = find_header_optional(headers, WH_CODE_CANDIDATES)
    status_idx = find_header_optional(headers, STATUS_CANDIDATES)

    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        wave_no = str(row[wave_idx] or "").strip()
        pick_time = parse_datetime(row[pick_time_idx])
        create_time = parse_datetime(row[create_time_idx]) if create_time_idx is not None else None
        ship_time = parse_datetime(row[ship_time_idx]) if ship_time_idx is not None else None
        order_no = str(row[order_no_idx] or "").strip() if order_no_idx is not None else ""
        customer = str(row[customer_idx] or "").strip() if customer_idx is not None else ""
        wh = str(row[wh_idx] or "").strip() if wh_idx is not None else ""
        status = str(row[status_idx] or "").strip() if status_idx is not None else ""

        # 计算各阶段耗时
        create_to_pick: float | None = None
        pick_to_ship: float | None = None
        create_to_ship: float | None = None

        if create_time and pick_time:
            create_to_pick = hours_between(create_time, pick_time)
        if pick_time and ship_time:
            pick_to_ship = hours_between(pick_time, ship_time)
        if create_time and ship_time:
            create_to_ship = hours_between(create_time, ship_time)

        orders.append({
            "order_no": order_no,
            "wave_no": wave_no,
            "customer": customer,
            "wh": wh,
            "status": status,
            "create_time": create_time,
            "pick_time": pick_time,
            "ship_time": ship_time,
            "create_to_pick": create_to_pick,
            "pick_to_ship": pick_to_ship,
            "create_to_ship": create_to_ship,
        })

    return orders


# ── 时效汇总 ────────────────────────────────────────────────
def build_leadtime_summary(
    workbook: Any,
    orders: list[dict[str, Any]],
    kpi_day: date,
) -> dict[str, Any]:
    """生成时效汇总统计"""
    iso_year, iso_week, _ = kpi_day.isocalendar()
    report_date = format_date(kpi_day)
    report_week = f"{iso_year}-W{iso_week:02d}"
    report_month = kpi_day.replace(day=1).strftime("%Y/%m/%d")

    total = len(orders)
    has_create = sum(1 for o in orders if o["create_time"])
    has_ship = sum(1 for o in orders if o["ship_time"])

    c2p_values = [o["create_to_pick"] for o in orders if o["create_to_pick"] is not None]
    p2s_values = [o["pick_to_ship"] for o in orders if o["pick_to_ship"] is not None]
    c2s_values = [o["create_to_ship"] for o in orders if o["create_to_ship"] is not None]

    # SLA 计算
    sla_12h = sum(1 for v in c2s_values if v <= 12)
    sla_24h = sum(1 for v in c2s_values if v <= 24)
    sla_48h = sum(1 for v in c2s_values if v <= 48)
    sla_total = len(c2s_values) if c2s_values else 1

    summary = {
        "日期": report_date, "周": report_week, "月": report_month,
        "总订单数": total,
        "有创建时间订单数": has_create,
        "有发货时间订单数": has_ship,
        "平均下单→拣货(小时)": round2(sum(c2p_values) / len(c2p_values)) if c2p_values else "",
        "中位数下单→拣货(小时)": median(c2p_values) if c2p_values else "",
        "平均拣货→发货(小时)": round2(sum(p2s_values) / len(p2s_values)) if p2s_values else "",
        "中位数拣货→发货(小时)": median(p2s_values) if p2s_values else "",
        "平均下单→发货(小时)": round2(sum(c2s_values) / len(c2s_values)) if c2s_values else "",
        "中位数下单→发货(小时)": median(c2s_values) if c2s_values else "",
        "最快下单→发货(小时)": round2(min(c2s_values)) if c2s_values else "",
        "最慢下单→发货(小时)": round2(max(c2s_values)) if c2s_values else "",
        "12h出库率": round2(sla_12h / sla_total * 100) if c2s_values else 0,
        "24h出库率": round2(sla_24h / sla_total * 100) if c2s_values else 0,
        "48h出库率": round2(sla_48h / sla_total * 100) if c2s_values else 0,
    }

    # Sheet: 时效汇总
    if LEADTIME_SUMMARY in workbook.sheetnames:
        del workbook[LEADTIME_SUMMARY]
    ws = workbook.create_sheet(LEADTIME_SUMMARY)
    ws.append(SUMMARY_HEADERS)
    ws.append([summary[h] for h in SUMMARY_HEADERS])
    style_sheet(ws)
    auto_width(ws)

    return summary


# ── 按波次时效 ──────────────────────────────────────────────
def build_wave_leadtime(workbook: Any, orders: list[dict[str, Any]]) -> None:
    """按波次号聚合时效"""
    wave_data: dict[str, dict[str, Any]] = {}
    for o in orders:
        wave = o["wave_no"]
        if not wave:
            continue
        if wave not in wave_data:
            wave_data[wave] = {"count": 0, "c2s": []}
        wave_data[wave]["count"] += 1
        if o["create_to_ship"] is not None:
            wave_data[wave]["c2s"].append(o["create_to_ship"])

    if WAVE_HEADERS[0] in workbook.sheetnames:
        del workbook[WAVE_HEADERS[0]]
    ws = workbook.create_sheet(LEADTIME_BY_WAVE)
    ws.append(WAVE_HEADERS)

    for wave, data in sorted(wave_data.items()):
        c2s = data["c2s"]
        avg_c2s = round2(sum(c2s) / len(c2s)) if c2s else ""
        min_c2s = round2(min(c2s)) if c2s else ""
        max_c2s = round2(max(c2s)) if c2s else ""
        ws.append([wave, data["count"], "", "", avg_c2s, min_c2s, max_c2s])

    style_sheet(ws)
    auto_width(ws)


# ── 时效分布 ────────────────────────────────────────────────
def build_distribution(workbook: Any, orders: list[dict[str, Any]]) -> None:
    """生成时效区间分布"""
    buckets = [
        ("≤2h", 0, 2),
        ("2-4h", 2, 4),
        ("4-8h", 4, 8),
        ("8-12h", 8, 12),
        ("12-24h", 12, 24),
        ("24-48h", 24, 48),
        ("48-72h", 48, 72),
        (">72h", 72, float("inf")),
    ]

    c2s_values = [o["create_to_ship"] for o in orders if o["create_to_ship"] is not None]
    total = len(c2s_values) if c2s_values else 1

    if DISTRIBUTION_SHEET in workbook.sheetnames:
        del workbook[DISTRIBUTION_SHEET]
    ws = workbook.create_sheet(DISTRIBUTION_SHEET)
    ws.append(DISTRIBUTION_HEADERS)

    for label, low, high in buckets:
        count = sum(1 for v in c2s_values if low < v <= high)
        pct = round2(count / total * 100)
        ws.append([label, count, f"{pct}%"])

    style_sheet(ws)
    auto_width(ws)


# ── SLA 达成率 ──────────────────────────────────────────────
def build_sla_sheet(workbook: Any, orders: list[dict[str, Any]]) -> None:
    """生成 SLA 达标率 sheet"""
    c2s_values = [o["create_to_ship"] for o in orders if o["create_to_ship"] is not None]
    total = len(c2s_values) if c2s_values else 1

    sla_levels = [
        ("12小时内出库", 12),
        ("24小时内出库", 24),
        ("48小时内出库", 48),
        ("72小时内出库", 72),
    ]

    if SLA_SHEET in workbook.sheetnames:
        del workbook[SLA_SHEET]
    ws = workbook.create_sheet(SLA_SHEET)
    ws.append(SLA_HEADERS)

    for label, hours in sla_levels:
        met = sum(1 for v in c2s_values if v <= hours)
        rate = round2(met / total * 100)
        ws.append([label, met, len(c2s_values), f"{rate}%"])

    # 条件格式：达标率<80% 红色，80-95% 黄色，>95% 绿色
    for row_idx in range(2, len(sla_levels) + 2):
        rate_cell = ws.cell(row=row_idx, column=4)
        try:
            rate_val = float(str(rate_cell.value).rstrip("%"))
            if rate_val < 80:
                rate_cell.fill = RED_FILL
            elif rate_val < 95:
                rate_cell.fill = YELLOW_FILL
            else:
                rate_cell.fill = GREEN_FILL
        except (ValueError, AttributeError):
            pass

    style_sheet(ws)
    auto_width(ws)


# ── 主流程 ──────────────────────────────────────────────────
def clean_leadtime(
    input_path: Path,
    output_path: Path,
    kpi_day: date,
) -> None:
    workbook = load_workbook(input_path)

    if MAIN_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"Excel cannot find main sheet: {MAIN_SHEET}")

    orders = extract_order_timeline(workbook)

    summary = build_leadtime_summary(workbook, orders, kpi_day)
    build_wave_leadtime(workbook, orders)
    build_distribution(workbook, orders)
    build_sla_sheet(workbook, orders)

    # 删除不需要的 sheet
    for sheet_name in list(workbook.sheetnames):
        if sheet_name not in (MAIN_SHEET, LEADTIME_SUMMARY, LEADTIME_BY_WAVE, DISTRIBUTION_SHEET, SLA_SHEET):
            del workbook[sheet_name]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"\n{'='*60}")
    print(f"KPI date: {kpi_day:%Y-%m-%d}")
    print(f"Total orders: {summary['总订单数']}")
    print(f"Orders with create time: {summary['有创建时间订单数']}")
    print(f"Orders with ship time: {summary['有发货时间订单数']}")
    print(f"Avg create→pick: {summary['平均下单→拣货(小时)']}h")
    print(f"Avg pick→ship: {summary['平均拣货→发货(小时)']}h")
    print(f"Avg create→ship: {summary['平均下单→发货(小时)']}h")
    print(f"12h SLA: {summary['12h出库率']}%")
    print(f"24h SLA: {summary['24h出库率']}%")
    print(f"48h SLA: {summary['48h出库率']}%")
    print(f"\nOutput: {output_path.resolve()}")
    print(f"{'='*60}")


def main() -> None:
    parser = argparse.ArgumentParser(description="出库时效 KPI 分析")
    parser.add_argument("input", help="Raw delivery export Excel path")
    parser.add_argument("--kpi-date", required=True, help="KPI date, YYYY-MM-DD")
    parser.add_argument("--output", default="", help="Output Excel path")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if args.output:
        output_path = Path(args.output).resolve()
    else:
        output_dir = input_path.parent / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{input_path.stem}_leadtime_kpi.xlsx"

    kpi_day = datetime.strptime(args.kpi_date, "%Y-%m-%d").date()
    clean_leadtime(input_path, output_path, kpi_day)


if __name__ == "__main__":
    main()
