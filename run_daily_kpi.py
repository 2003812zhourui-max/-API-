from __future__ import annotations

import argparse
from argparse import Namespace
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from clean_delivery_kpi import clean_workbook as clean_pick_workbook
from clean_delivery_kpi import parse_kpi_date
from clean_review_kpi import clean_review_workbook


PICK_PERSON_SHEET = "拣货人汇总"
REVIEW_PERSON_SHEET = "复核人员汇总"
UNIFIED_PERSON_SHEET = "人员KPI汇总"

UNIFIED_HEADERS = [
    "日期",
    "周",
    "月",
    "环节",
    "名字",
    "总数量",
    "订单数",
    "波次数量",
    "最早时间",
    "最晚时间",
    "工作小时数",
    "工作时间",
    "总数占比",
    "波次占比",
    "件/小时",
    "波次/小时",
    "平均每波总数",
    "平均每波订单数",
    "平均每波工作分钟",
]


def compact_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def header_index(headers: list[Any], name: str) -> int:
    target = compact_header(name)
    for index, header in enumerate(headers):
        if compact_header(header) == target:
            return index
    raise RuntimeError(f"Missing column: {name}")


def cell(row: tuple[Any, ...], indexes: dict[str, int], name: str, default: Any = "") -> Any:
    index = indexes.get(name)
    if index is None or index >= len(row):
        return default
    value = row[index]
    return default if value is None else value


def style_header(sheet: Any) -> None:
    for item in sheet[1]:
        item.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def resize_columns(sheet: Any, max_width: int = 32) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 10
        for item in column_cells:
            value = "" if item.value is None else str(item.value)
            width = max(width, min(len(value) + 2, max_width))
        sheet.column_dimensions[column_letter].width = width


def read_pick_person_rows(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if PICK_PERSON_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"{path} missing sheet: {PICK_PERSON_SHEET}")
    sheet = workbook[PICK_PERSON_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [compact_header(value) for value in rows[0]]
    indexes = {header: index for index, header in enumerate(headers) if header}
    result: list[list[Any]] = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue
        result.append(
            [
                cell(row, indexes, "日期"),
                cell(row, indexes, "周"),
                cell(row, indexes, "月"),
                "拣货",
                cell(row, indexes, "名字", "未匹配"),
                cell(row, indexes, "总数"),
                cell(row, indexes, "重复次数"),
                cell(row, indexes, "波次数量"),
                cell(row, indexes, "最早拣货时间"),
                cell(row, indexes, "最晚拣货时间"),
                cell(row, indexes, "工作小时数"),
                cell(row, indexes, "工作时间"),
                cell(row, indexes, "总数占比"),
                cell(row, indexes, "波次占比"),
                cell(row, indexes, "件/小时"),
                cell(row, indexes, "波次/小时"),
                cell(row, indexes, "平均每波总数"),
                cell(row, indexes, "平均每波重复次数"),
                cell(row, indexes, "平均每波工作分钟"),
            ]
        )
    return result


def read_review_person_rows(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if REVIEW_PERSON_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"{path} missing sheet: {REVIEW_PERSON_SHEET}")
    sheet = workbook[REVIEW_PERSON_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [compact_header(value) for value in rows[0]]
    indexes = {header: index for index, header in enumerate(headers) if header}
    result: list[list[Any]] = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue
        result.append(
            [
                cell(row, indexes, "日期"),
                cell(row, indexes, "周"),
                cell(row, indexes, "月"),
                cell(row, indexes, "环节", "复核-一单一件"),
                cell(row, indexes, "名字", "未匹配"),
                cell(row, indexes, "总数量"),
                cell(row, indexes, "订单数"),
                cell(row, indexes, "波次数量"),
                cell(row, indexes, "最早复核时间"),
                cell(row, indexes, "最晚复核时间"),
                cell(row, indexes, "工作小时数"),
                cell(row, indexes, "工作时间"),
                cell(row, indexes, "总数占比"),
                cell(row, indexes, "波次占比"),
                cell(row, indexes, "件/小时"),
                cell(row, indexes, "波次/小时"),
                cell(row, indexes, "平均每波总数"),
                cell(row, indexes, "平均每波订单数"),
                cell(row, indexes, "平均每波工作分钟"),
            ]
        )
    return result


def build_unified_person_workbook(pick_path: Path, review_path: Path, output_path: Path) -> int:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = UNIFIED_PERSON_SHEET
    sheet.append(UNIFIED_HEADERS)

    rows = read_pick_person_rows(pick_path) + read_review_person_rows(review_path)
    for row in rows:
        sheet.append(row)

    style_header(sheet)
    resize_columns(sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(rows)


def workbook_has_sheet(path: Path, sheet_name: str) -> bool:
    workbook = load_workbook(path, read_only=True, data_only=True)
    return sheet_name in workbook.sheetnames


def default_output_dir(kpi_day: date) -> Path:
    return Path.cwd() / "output" / "kpi" / kpi_day.strftime("%Y%m%d")


def build_pick_args(args: argparse.Namespace) -> Namespace:
    return Namespace(
        name=args.name,
        fetch_picker_names=True,
        storage_state=args.storage_state,
        picker_start_time=args.picker_start_time,
        picker_end_time=args.picker_end_time,
        picker_page_size=args.picker_page_size,
        picker_sample_limit=args.picker_sample_limit,
        picker_wave_limit=args.picker_wave_limit,
        picker_sleep=args.picker_sleep,
    )


def build_review_args(args: argparse.Namespace) -> Namespace:
    return Namespace(
        name=args.name,
        fetch_reviewer_names=True,
        storage_state=args.storage_state,
        order_start_time=args.order_start_time,
        order_end_time=args.order_end_time,
        reviewer_page_size=args.reviewer_page_size,
        reviewer_sample_limit=args.reviewer_sample_limit,
        reviewer_wave_limit=args.reviewer_wave_limit,
        reviewer_sleep=args.reviewer_sleep,
    )


def run_daily_kpi(args: argparse.Namespace) -> None:
    kpi_day = parse_kpi_date(args.kpi_date)
    input_path = Path(args.input).resolve()
    storage_state = Path(args.storage_state).resolve()
    if not storage_state.exists():
        raise FileNotFoundError(f"storage_state not found: {storage_state}")
    args.storage_state = str(storage_state)

    output_dir = Path(args.output_dir).resolve() if args.output_dir else default_output_dir(kpi_day).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    pick_path = output_dir / f"pick_kpi_{kpi_day:%Y%m%d}.xlsx"
    review_path = output_dir / f"review_kpi_{kpi_day:%Y%m%d}.xlsx"
    unified_path = output_dir / f"person_kpi_summary_{kpi_day:%Y%m%d}.xlsx"

    print("=== 1/3 拣货 KPI ===", flush=True)
    if workbook_has_sheet(input_path, PICK_PERSON_SHEET):
        pick_summary_source = input_path
        print(f"reuse existing pick summary sheet: {input_path} / {PICK_PERSON_SHEET}", flush=True)
    else:
        pick_summary_source = pick_path
        clean_pick_workbook(input_path, pick_path, kpi_day, build_pick_args(args))

    print("=== 2/3 复核 KPI ===", flush=True)
    clean_review_workbook(input_path, review_path, kpi_day, build_review_args(args))

    print("=== 3/3 统一人员汇总 ===", flush=True)
    summary_rows = build_unified_person_workbook(pick_summary_source, review_path, unified_path)

    print(f"pick source: {pick_summary_source}")
    print(f"review file: {review_path}")
    print(f"summary rows: {summary_rows}")
    print(f"summary file: {unified_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Run daily pick + review KPI with OMP name lookup.")
    parser.add_argument("input", help="Delivery Excel file path")
    parser.add_argument("--kpi-date", required=True, help="KPI date, YYYY-MM-DD")
    parser.add_argument("--storage-state", default="storage_state.json")
    parser.add_argument("--output-dir", default="")
    parser.add_argument("--name", default="", help="Fallback name when name lookup fails")

    parser.add_argument("--picker-start-time", default="")
    parser.add_argument("--picker-end-time", default="")
    parser.add_argument("--picker-page-size", type=int, default=20)
    parser.add_argument("--picker-sample-limit", type=int, default=3)
    parser.add_argument("--picker-wave-limit", type=int, default=0)
    parser.add_argument("--picker-sleep", type=float, default=0.0)

    parser.add_argument("--order-start-time", default="")
    parser.add_argument("--order-end-time", default="")
    parser.add_argument("--reviewer-page-size", type=int, default=100)
    parser.add_argument("--reviewer-sample-limit", type=int, default=3)
    parser.add_argument("--reviewer-wave-limit", type=int, default=0)
    parser.add_argument("--reviewer-sleep", type=float, default=0.0)
    args = parser.parse_args()

    run_daily_kpi(args)


if __name__ == "__main__":
    main()
