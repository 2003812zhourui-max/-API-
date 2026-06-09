from __future__ import annotations

import argparse
import time as time_module
from collections import Counter, OrderedDict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from clean_delivery_kpi import (
    API_ORIGIN,
    compact_json_text,
    create_omp_session,
    delivery_page_payload,
    format_date,
    parse_datetime,
    parse_kpi_date,
    parse_number,
)
from track_key import track_key_for_text


MAIN_SHEET = "出库单"
FILTERED_SHEET = "出库时间窗口"
WAVE_SUMMARY_SHEET = "出库波次号去重统计"
PERSON_SUMMARY_SHEET = "出库人员汇总"

ORDER_NO_HEADER = "Outbound Order No/出库单号"
WAVE_NO_HEADER = "Wave No./关联波次号"
REVIEW_TIME_HEADER = "Review time/复核时间"
OUTBOUND_TIME_HEADER = "OutboundTime/出库时间"
TOTAL_QTY_HEADER = "Total Qty of SKU/总数量"

WAVE_SUMMARY_HEADERS = [
    "日期",
    "周",
    "月",
    "名字",
    "波次号",
    "总数量",
    "订单数",
    "最早出库时间",
    "最晚出库时间",
]

PERSON_SUMMARY_HEADERS = [
    "日期",
    "周",
    "月",
    "环节",
    "名字",
    "总数量",
    "订单数",
    "最早出库时间",
    "最晚出库时间",
    "波次数量",
    "工作小时数",
    "工作时间",
    "总数占比",
    "波次占比",
    "件/小时",
    "波次/小时",
    "平均每波总数",
    "平均每波订单数",
    "平均每波工作分钟",
]


def compact_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def find_header(headers: list[Any], target: str) -> int:
    normalized_target = compact_header(target)
    for index, header in enumerate(headers):
        if compact_header(header) == normalized_target:
            return index
    raise RuntimeError(f"Excel missing required column: {target}")


def reset_sheet(workbook: Any, title: str) -> None:
    if title in workbook.sheetnames:
        workbook.remove(workbook[title])


def style_header(sheet: Any) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def resize_columns(sheet: Any, max_width: int = 32) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, max_width))
        sheet.column_dimensions[column_letter].width = width


def format_datetime(value: datetime | None) -> str:
    return "" if value is None else value.strftime("%Y/%m/%d %H:%M:%S")


def format_hours(hours: float) -> str:
    total_minutes = int(round(hours * 60))
    return f"{total_minutes // 60}小时{total_minutes % 60:02d}分"


def outbound_time_window(kpi_day: date) -> tuple[datetime, datetime]:
    return datetime.combine(kpi_day - timedelta(days=1), time(22, 30, 0)), datetime.combine(kpi_day, time(11, 0, 0))


def create_time_window(kpi_day: date) -> tuple[datetime, datetime]:
    if kpi_day.weekday() == 0:
        return datetime.combine(kpi_day - timedelta(days=2), time.min), datetime.combine(
            kpi_day - timedelta(days=1), time(22, 30, 0)
        )
    if kpi_day.weekday() == 5:
        return datetime.combine(kpi_day - timedelta(days=2), time(22, 30, 0)), datetime.combine(
            kpi_day - timedelta(days=1), time(23, 59, 59)
        )
    return datetime.combine(kpi_day - timedelta(days=1), time(22, 30, 0)), datetime.combine(
        kpi_day, time(22, 30, 0)
    )


def remove_unwanted_sheets(workbook: Any) -> None:
    if MAIN_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"Excel cannot find main sheet: {MAIN_SHEET}")
    for sheet in list(workbook.worksheets):
        if sheet.title != MAIN_SHEET:
            workbook.remove(sheet)


def is_outbound_kpi_row(
    row: tuple[Any, ...],
    qty_index: int,
    review_index: int,
    outbound_index: int,
    start_at: datetime,
    end_at: datetime,
) -> bool:
    qty = parse_number(row[qty_index])
    review_time = parse_datetime(row[review_index])
    outbound_time = parse_datetime(row[outbound_index])
    return (
        qty is not None
        and qty != 1
        and review_time is None
        and outbound_time is not None
        and start_at <= outbound_time <= end_at
    )


def build_filtered_sheet(workbook: Any, kpi_day: date) -> tuple[int, int]:
    source = workbook[MAIN_SHEET]
    headers = [cell.value for cell in source[1]]
    qty_index = find_header(headers, TOTAL_QTY_HEADER)
    review_index = find_header(headers, REVIEW_TIME_HEADER)
    outbound_index = find_header(headers, OUTBOUND_TIME_HEADER)
    start_at, end_at = outbound_time_window(kpi_day)

    reset_sheet(workbook, FILTERED_SHEET)
    filtered = workbook.create_sheet(FILTERED_SHEET)
    filtered.append(headers)

    kept_rows = 0
    removed_rows = 0
    for row in source.iter_rows(min_row=2, values_only=True):
        if is_outbound_kpi_row(row, qty_index, review_index, outbound_index, start_at, end_at):
            filtered.append(list(row))
            kept_rows += 1
        else:
            removed_rows += 1

    style_header(filtered)
    resize_columns(filtered)
    return kept_rows, removed_rows


def collect_wave_groups(workbook: Any) -> OrderedDict[str, dict[str, Any]]:
    filtered = workbook[FILTERED_SHEET]
    headers = [cell.value for cell in filtered[1]]
    order_index = find_header(headers, ORDER_NO_HEADER)
    wave_index = find_header(headers, WAVE_NO_HEADER)
    outbound_index = find_header(headers, OUTBOUND_TIME_HEADER)
    qty_index = find_header(headers, TOTAL_QTY_HEADER)

    groups: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row in filtered.iter_rows(min_row=2, values_only=True):
        wave_no = str(row[wave_index] or "").strip()
        if not wave_no:
            continue
        group = groups.setdefault(
            wave_no,
            {"order_nos": [], "qty": 0.0, "has_qty": False, "count": 0, "min_time": None, "max_time": None},
        )
        order_no = str(row[order_index] or "").strip()
        if order_no:
            group["order_nos"].append(order_no)
        group["count"] += 1

        qty = parse_number(row[qty_index])
        if qty is not None:
            group["qty"] += qty
            group["has_qty"] = True

        outbound_time = parse_datetime(row[outbound_index])
        if outbound_time is not None:
            group["min_time"] = outbound_time if group["min_time"] is None else min(group["min_time"], outbound_time)
            group["max_time"] = outbound_time if group["max_time"] is None else max(group["max_time"], outbound_time)
    return groups


def fetch_delivery_records_by_wave(
    session: requests.Session,
    wave_no: str,
    start_time: str,
    end_time: str,
    page_size: int,
) -> list[dict[str, Any]]:
    payload = delivery_page_payload(wave_no, start_time, end_time, page_size)
    body = compact_json_text(payload)
    try:
        track_key = track_key_for_text(body)
    except Exception:
        track_key = ""
    if not track_key:
        raise RuntimeError("Could not generate Track-Key for delivery page request.")

    response = session.post(
        f"{API_ORIGIN}/gateway/omp/order/delivery/page",
        data=body.encode("utf-8"),
        headers={"Content-Type": "application/json;charset=UTF-8", "Track-Key": track_key},
        timeout=60,
    )
    response.raise_for_status()
    result = response.json()
    if result.get("code") != 200:
        raise RuntimeError(f"delivery page failed for {wave_no}: {result}")
    records = (result.get("data") or {}).get("records") or []
    return records if isinstance(records, list) else []


def looks_like_outbound_event(event: dict[str, Any]) -> bool:
    biz_type = str(event.get("bizType") or "").strip().lower()
    biz_desc = str(event.get("bizDesc") or "")
    return biz_type == "delivery_outbound" or "出库" in biz_desc


def fetch_outbound_name_from_order(session: requests.Session, record: dict[str, Any]) -> str:
    delivery_no = str(record.get("deliveryNo") or "").strip()
    customer_code = str(record.get("customerCode") or "").strip()
    wh_code = str(record.get("whCode") or "").strip()
    source_no = str(record.get("sourceNo") or "").strip()
    if not delivery_no or not customer_code or not wh_code:
        return ""

    query = urlencode({"billType": "delivery_packet", "billNo": delivery_no, "customerCode": customer_code, "whCode": wh_code})
    response = session.get(
        f"{API_ORIGIN}/gateway/omp/order/log/progress?{query}",
        headers={"Referer": f"{API_ORIGIN}/globalOrder/parcelDetail/{customer_code}/{wh_code}/{delivery_no}/{source_no}"},
        timeout=60,
    )
    response.raise_for_status()
    result = response.json()
    if result.get("code") != 200:
        return ""
    for event in result.get("data") or []:
        if looks_like_outbound_event(event):
            return str(event.get("operateAccount") or event.get("operateName") or "").strip()
    return ""


def resolve_outbound_names(
    wave_groups: OrderedDict[str, dict[str, Any]],
    storage_state: str,
    start_time: str,
    end_time: str,
    page_size: int,
    sample_limit: int,
    wave_limit: int,
    sleep_seconds: float,
) -> dict[str, str]:
    session = create_omp_session(storage_state)
    result: dict[str, str] = {}
    target_items = list(wave_groups.items())
    if wave_limit > 0:
        target_items = target_items[:wave_limit]

    for index, (wave_no, group) in enumerate(target_items, 1):
        try:
            records = fetch_delivery_records_by_wave(session, wave_no, start_time, end_time, page_size)
            target_order_nos = set(group["order_nos"])
            sample_records = [
                record
                for record in records
                if str(record.get("sourceNo") or record.get("deliveryNo") or "").strip() in target_order_nos
            ] or records

            names: list[str] = []
            for record in sample_records[:sample_limit]:
                name = fetch_outbound_name_from_order(session, record)
                if name:
                    names.append(name)
            if names:
                result[wave_no] = Counter(names).most_common(1)[0][0]
            print(f"outbound {index}/{len(target_items)} {wave_no}: {result.get(wave_no, '') or '-'}", flush=True)
        except Exception as exc:
            result[wave_no] = ""
            print(f"outbound {index}/{len(target_items)} {wave_no}: failed: {exc}", flush=True)
        if sleep_seconds > 0:
            time_module.sleep(sleep_seconds)
    return result


def report_dimensions(kpi_day: date) -> tuple[str, str, str]:
    iso_year, iso_week, _ = kpi_day.isocalendar()
    return format_date(kpi_day), f"{iso_year}-W{iso_week:02d}", kpi_day.replace(day=1).strftime("%Y/%m/%d")


def numeric_qty(value: dict[str, Any]) -> int | float | str:
    if not value["has_qty"]:
        return ""
    qty = value["qty"]
    return int(qty) if qty.is_integer() else round(qty, 2)


def build_wave_summary_sheet(workbook: Any, kpi_day: date, fallback_name: str, outbound_names: dict[str, str]) -> int:
    groups = collect_wave_groups(workbook)
    reset_sheet(workbook, WAVE_SUMMARY_SHEET)
    summary = workbook.create_sheet(WAVE_SUMMARY_SHEET)
    summary.append(WAVE_SUMMARY_HEADERS)

    report_date, report_week, report_month = report_dimensions(kpi_day)
    for wave_no, group in groups.items():
        summary.append(
            [
                report_date,
                report_week,
                report_month,
                outbound_names.get(wave_no) or fallback_name,
                wave_no,
                numeric_qty(group),
                group["count"],
                format_datetime(group["min_time"]),
                format_datetime(group["max_time"]),
            ]
        )

    style_header(summary)
    resize_columns(summary)
    return len(groups)


def build_person_summary_sheet(workbook: Any, kpi_day: date) -> int:
    wave_summary = workbook[WAVE_SUMMARY_SHEET]
    headers = [cell.value for cell in wave_summary[1]]
    name_index = find_header(headers, "名字")
    qty_index = find_header(headers, "总数量")
    count_index = find_header(headers, "订单数")
    min_index = find_header(headers, "最早出库时间")
    max_index = find_header(headers, "最晚出库时间")

    people: OrderedDict[str, dict[str, Any]] = OrderedDict()
    total_qty = 0.0
    total_waves = 0
    for row in wave_summary.iter_rows(min_row=2, values_only=True):
        name = str(row[name_index] or "").strip() or "未匹配"
        person = people.setdefault(name, {"qty": 0.0, "orders": 0, "waves": 0, "min_time": None, "max_time": None})
        qty = parse_number(row[qty_index]) or 0.0
        orders = int(parse_number(row[count_index]) or 0)
        min_time = parse_datetime(row[min_index])
        max_time = parse_datetime(row[max_index])

        person["qty"] += qty
        person["orders"] += orders
        person["waves"] += 1
        if min_time is not None:
            person["min_time"] = min_time if person["min_time"] is None else min(person["min_time"], min_time)
        if max_time is not None:
            person["max_time"] = max_time if person["max_time"] is None else max(person["max_time"], max_time)
        total_qty += qty
        total_waves += 1

    reset_sheet(workbook, PERSON_SUMMARY_SHEET)
    summary = workbook.create_sheet(PERSON_SUMMARY_SHEET)
    summary.append(PERSON_SUMMARY_HEADERS)
    report_date, report_week, report_month = report_dimensions(kpi_day)

    for name, person in people.items():
        min_time = person["min_time"]
        max_time = person["max_time"]
        hours = 0.0
        if min_time is not None and max_time is not None and max_time > min_time:
            hours = round((max_time - min_time).total_seconds() / 3600, 6)
        qty = person["qty"]
        waves = person["waves"]
        orders = person["orders"]

        summary.append(
            [
                report_date,
                report_week,
                report_month,
                "出库-一单多件",
                name,
                int(qty) if qty.is_integer() else round(qty, 2),
                orders,
                format_datetime(min_time),
                format_datetime(max_time),
                waves,
                hours,
                format_hours(hours),
                round(qty / total_qty, 6) if total_qty else "",
                round(waves / total_waves, 6) if total_waves else "",
                round(qty / hours, 2) if hours else "",
                round(waves / hours, 2) if hours else "",
                round(qty / waves, 2) if waves else "",
                round(orders / waves, 2) if waves else "",
                round(hours * 60 / waves, 2) if waves else "",
            ]
        )

    style_header(summary)
    resize_columns(summary)
    return len(people)


def default_output_path(input_path: Path, kpi_day: date) -> Path:
    output_dir = input_path.parent / "output"
    if input_path.parent == Path.cwd():
        output_dir = Path.cwd() / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"{input_path.stem}_outbound_kpi_{kpi_day:%Y%m%d}.xlsx"


def clean_outbound_workbook(input_path: Path, output_path: Path, kpi_day: date, args: argparse.Namespace) -> None:
    workbook = load_workbook(input_path)
    remove_unwanted_sheets(workbook)
    kept_rows, removed_rows = build_filtered_sheet(workbook, kpi_day)
    groups = collect_wave_groups(workbook)

    outbound_names: dict[str, str] = {}
    if args.fetch_outbound_names:
        create_start, create_end = create_time_window(kpi_day)
        start_time = args.order_start_time or create_start.strftime("%Y-%m-%d %H:%M:%S")
        end_time = args.order_end_time or create_end.strftime("%Y-%m-%d %H:%M:%S")
        outbound_names = resolve_outbound_names(
            wave_groups=groups,
            storage_state=args.storage_state,
            start_time=start_time,
            end_time=end_time,
            page_size=args.outbound_page_size,
            sample_limit=args.outbound_sample_limit,
            wave_limit=args.outbound_wave_limit,
            sleep_seconds=args.outbound_sleep,
        )

    wave_count = build_wave_summary_sheet(workbook, kpi_day, args.name, outbound_names)
    person_count = build_person_summary_sheet(workbook, kpi_day)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    outbound_start, outbound_end = outbound_time_window(kpi_day)
    create_start, create_end = create_time_window(kpi_day)
    print(f"KPI date: {kpi_day:%Y-%m-%d}")
    print(f"create window: {create_start:%Y-%m-%d %H:%M:%S} ~ {create_end:%Y-%m-%d %H:%M:%S}")
    print(f"outbound window: {outbound_start:%Y-%m-%d %H:%M:%S} ~ {outbound_end:%Y-%m-%d %H:%M:%S}")
    print("filter: Total Qty of SKU != 1, Review time blank, OutboundTime within window")
    print(f"kept rows: {kept_rows}")
    print(f"removed rows: {removed_rows}")
    print(f"outbound wave rows: {wave_count}")
    print(f"outbound person rows: {person_count}")
    if args.fetch_outbound_names:
        print(f"outbound names found: {sum(1 for value in outbound_names.values() if value)}")
    print(f"output file: {output_path.resolve()}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build outbound KPI workbook from OMP delivery export Excel.")
    parser.add_argument("input", help="Raw or cleaned delivery Excel file path")
    parser.add_argument("--kpi-date", required=True)
    parser.add_argument("--name", default="未匹配")
    parser.add_argument("--output", default="")
    parser.add_argument("--fetch-outbound-names", action="store_true")
    parser.add_argument("--storage-state", default="")
    parser.add_argument("--order-start-time", default="")
    parser.add_argument("--order-end-time", default="")
    parser.add_argument("--outbound-page-size", type=int, default=100)
    parser.add_argument("--outbound-sample-limit", type=int, default=3)
    parser.add_argument("--outbound-wave-limit", type=int, default=0)
    parser.add_argument("--outbound-sleep", type=float, default=0.0)
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    output_path = Path(args.output).resolve() if args.output else default_output_path(input_path, parse_kpi_date(args.kpi_date)).resolve()
    clean_outbound_workbook(input_path, output_path, parse_kpi_date(args.kpi_date), args)


if __name__ == "__main__":
    main()
