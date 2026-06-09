from __future__ import annotations

import argparse
import json
import os
import time as time_module
from collections import Counter, OrderedDict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from track_key import track_key_for_text


API_ORIGIN = "https://omp.xlwms.com"

MAIN_SHEET = "\u51fa\u5e93\u5355"
FILTERED_SHEET = "\u62e3\u8d27\u65f6\u95f4\u53bb\u9664"
SUMMARY_SHEET = "\u6ce2\u6b21\u53f7\u53bb\u91cd\u7edf\u8ba1"

PICK_TIME_HEADER = "Pick time/\u62e3\u8d27\u65f6\u95f4"
WAVE_NO_HEADER = "Wave No./\u5173\u8054\u6ce2\u6b21\u53f7"
TOTAL_QTY_HEADER = "Total Qty of SKU/\u603b\u6570\u91cf"

SUMMARY_HEADERS = [
    "\u65e5\u671f",
    "\u5468",
    "\u6708",
    "\u540d\u5b57",
    "\u6ce2\u6b21\u53f7",
    "\u603b\u6570\u91cf",
    "\u6b21\u53f7\u91cd\u590d\u6b21",
    "\u6700\u65e9\u62e3\u8d27\u65f6\u95f4",
    "\u6700\u665a\u62e3\u8d27\u65f6\u95f4",
]


def compact_json_text(value: Any) -> str:
    # ensure_ascii=True required to match browser Track-Key signing
    return json.dumps(value, ensure_ascii=True, separators=(",", ":"))


def parse_kpi_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def compact_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def find_header(headers: list[Any], target: str) -> int:
    normalized_target = compact_header(target)
    for index, header in enumerate(headers):
        if compact_header(header) == normalized_target:
            return index
    raise RuntimeError(f"Excel missing required column: {target}")


def kpi_window(kpi_day: date) -> tuple[datetime, datetime]:
    start_at = datetime.combine(kpi_day - timedelta(days=1), time(22, 30, 0))
    if kpi_day.weekday() == 4:
        end_at = datetime.combine(kpi_day, time(23, 59, 59))
    else:
        end_at = datetime.combine(kpi_day, time(22, 30, 0))
    return start_at, end_at


def picker_search_window(kpi_day: date) -> tuple[datetime, datetime]:
    return kpi_window(kpi_day)


def pick_time_exclusion_window(kpi_day: date) -> tuple[datetime, datetime]:
    start_at = datetime.combine(kpi_day - timedelta(days=1), time(22, 30, 0))
    end_at = datetime.combine(kpi_day, time(10, 30, 0))
    return start_at, end_at


def reset_sheet(workbook: Any, title: str) -> None:
    if title in workbook.sheetnames:
        workbook.remove(workbook[title])


def style_header(sheet: Any) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def resize_columns(sheet: Any, max_width: int = 32) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, max_width))
        sheet.column_dimensions[column_letter].width = width


def format_date(value: date) -> str:
    return value.strftime("%Y/%m/%d")


def format_datetime(value: datetime | None) -> str:
    return "" if value is None else value.strftime("%Y/%m/%d %H:%M:%S")


def load_storage_state(path: str) -> dict[str, Any]:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def local_storage_from_state(state: dict[str, Any], origin: str = API_ORIGIN) -> dict[str, str]:
    result: dict[str, str] = {}
    for origin_entry in state.get("origins", []):
        if origin_entry.get("origin") != origin:
            continue
        for entry in origin_entry.get("localStorage", []):
            result[str(entry.get("name"))] = str(entry.get("value"))
    return result


def create_omp_session(storage_state: str = "") -> requests.Session:
    session = requests.Session()

    if storage_state:
        state = load_storage_state(storage_state)
        for cookie in state.get("cookies", []):
            session.cookies.set(
                cookie["name"],
                cookie["value"],
                domain=cookie.get("domain"),
                path=cookie.get("path", "/"),
            )
        local_storage = local_storage_from_state(state)
        token = local_storage.get("omp-token") or local_storage.get("wms-token") or ""
    else:
        token = ""

    env_token = os.environ.get("OMP_TOKEN", "").strip()
    token = env_token or token
    if token.lower().startswith("bearer "):
        token = token[7:].strip()
    if not token:
        raise RuntimeError("Set OMP_TOKEN or pass --storage-state before fetching picker names.")

    cookie = os.environ.get("OMP_COOKIE", "").strip()
    if cookie:
        session.headers["Cookie"] = cookie

    session.headers.update(
        {
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Authorization": f"Bearer {token}",
            "Origin": API_ORIGIN,
            "Referer": f"{API_ORIGIN}/globalOrder/parcel",
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
            ),
            "lang": "zh",
            "version": "prod",
        }
    )
    return session


def delivery_page_payload(
    wave_no: str,
    start_time: str,
    end_time: str,
    size: int,
    current: int = 1,
) -> dict[str, Any]:
    return {
        "appendixFlag": "",
        "categoryIdList": [],
        "countKind": "orderWeight",
        "current": current,
        "endTime": end_time,
        "expressFlag": "",
        "morePkgFlag": "",
        "noTypeNo": wave_no,
        "orderSourceList": [],
        "platformCode": 1,
        "relatedReturnOrder": "",
        "size": size,
        "startTime": start_time,
        "status": "",
        "storeName": "",
        "timeType": "createTime",
        "transitStatusList": [],
        "unitMark": 0,
        "waveNo": wave_no,
        "waveNoList": [wave_no],
        "whCode": "",
        "withVas": "",
    }


def fetch_delivery_records_by_wave(
    session: requests.Session,
    wave_no: str,
    start_time: str,
    end_time: str,
    page_size: int,
) -> list[dict[str, Any]]:
    payload = delivery_page_payload(wave_no, start_time, end_time, page_size)
    body = compact_json_text(payload)
    try:
        track_key = track_key_for_text(body)
    except Exception:
        track_key = os.environ.get("OMP_TRACK_KEY", "").strip()
    if not track_key:
        raise RuntimeError("Could not generate Track-Key for delivery page request.")

    response = session.post(
        f"{API_ORIGIN}/gateway/omp/order/delivery/page",
        data=body.encode("utf-8"),
        headers={
            "Content-Type": "application/json;charset=UTF-8",
            "Track-Key": track_key,
            "Referer": f"{API_ORIGIN}/globalOrder/parcel",
        },
        timeout=60,
    )
    response.raise_for_status()
    result = response.json()
    if result.get("code") != 200:
        raise RuntimeError(f"delivery page failed for {wave_no}: {result}")
    records = (result.get("data") or {}).get("records") or []
    return records if isinstance(records, list) else []


def fetch_pick_name_from_order(session: requests.Session, record: dict[str, Any]) -> str:
    delivery_no = str(record.get("deliveryNo") or "").strip()
    customer_code = str(record.get("customerCode") or "").strip()
    wh_code = str(record.get("whCode") or "").strip()
    source_no = str(record.get("sourceNo") or "").strip()
    if not delivery_no or not customer_code or not wh_code:
        return ""

    query = urlencode(
        {
            "billType": "delivery_packet",
            "billNo": delivery_no,
            "customerCode": customer_code,
            "whCode": wh_code,
        }
    )
    response = session.get(
        f"{API_ORIGIN}/gateway/omp/order/log/progress?{query}",
        headers={"Referer": f"{API_ORIGIN}/globalOrder/parcelDetail/{customer_code}/{wh_code}/{delivery_no}/{source_no}"},
        timeout=60,
    )
    response.raise_for_status()
    result = response.json()
    if result.get("code") != 200:
        return ""

    events = result.get("data") or []
    for event in events:
        if event.get("bizType") == "delivery_pick" or event.get("bizDesc") == "\u62e3\u8d27":
            return str(event.get("operateAccount") or event.get("operateName") or "").strip()
    return ""


def resolve_picker_names(
    wave_nos: list[str],
    storage_state: str,
    start_time: str,
    end_time: str,
    page_size: int,
    sample_limit: int,
    wave_limit: int,
    sleep_seconds: float,
) -> dict[str, str]:
    session = create_omp_session(storage_state)
    result: dict[str, str] = {}
    target_wave_nos = wave_nos[:wave_limit] if wave_limit > 0 else wave_nos

    for index, wave_no in enumerate(target_wave_nos, 1):
        if not wave_no:
            continue
        try:
            records = fetch_delivery_records_by_wave(session, wave_no, start_time, end_time, page_size)
            names: list[str] = []
            for record in records[:sample_limit]:
                name = fetch_pick_name_from_order(session, record)
                if name:
                    names.append(name)
            if names:
                result[wave_no] = Counter(names).most_common(1)[0][0]
            print(f"picker {index}/{len(target_wave_nos)} {wave_no}: {result.get(wave_no, '') or '-'}", flush=True)
        except Exception as exc:
            result[wave_no] = ""
            print(f"picker {index}/{len(target_wave_nos)} {wave_no}: failed: {exc}", flush=True)

        if sleep_seconds > 0:
            time_module.sleep(sleep_seconds)

    return result


def build_filtered_sheet(workbook: Any, kpi_day: date) -> tuple[int, int]:
    source = workbook[MAIN_SHEET]
    headers = [cell.value for cell in source[1]]
    pick_index = find_header(headers, PICK_TIME_HEADER)
    exclude_start, exclude_end = pick_time_exclusion_window(kpi_day)

    reset_sheet(workbook, FILTERED_SHEET)
    filtered = workbook.create_sheet(FILTERED_SHEET)
    filtered.append(headers)

    kept_rows = 0
    removed_rows = 0
    for row in source.iter_rows(min_row=2, values_only=True):
        pick_time = parse_datetime(row[pick_index])
        should_remove = pick_time is not None and exclude_start <= pick_time <= exclude_end
        if should_remove:
            removed_rows += 1
            continue
        filtered.append(list(row))
        kept_rows += 1

    style_header(filtered)
    resize_columns(filtered)
    return kept_rows, removed_rows


def collect_wave_groups(workbook: Any) -> OrderedDict[str, dict[str, Any]]:
    filtered = workbook[FILTERED_SHEET]
    headers = [cell.value for cell in filtered[1]]
    wave_index = find_header(headers, WAVE_NO_HEADER)
    pick_index = find_header(headers, PICK_TIME_HEADER)
    qty_index = find_header(headers, TOTAL_QTY_HEADER)

    groups: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row in filtered.iter_rows(min_row=2, values_only=True):
        wave_no = str(row[wave_index] or "").strip()
        if not wave_no:
            continue

        group = groups.setdefault(
            wave_no,
            {
                "count": 0,
                "qty": 0.0,
                "has_qty": False,
                "min_pick": None,
                "max_pick": None,
            },
        )
        group["count"] += 1

        qty = parse_number(row[qty_index])
        if qty is not None:
            group["qty"] += qty
            group["has_qty"] = True

        pick_time = parse_datetime(row[pick_index])
        if pick_time is not None:
            group["min_pick"] = pick_time if group["min_pick"] is None else min(group["min_pick"], pick_time)
            group["max_pick"] = pick_time if group["max_pick"] is None else max(group["max_pick"], pick_time)
    return groups


def build_summary_sheet(
    workbook: Any,
    kpi_day: date,
    fallback_name: str,
    picker_names: dict[str, str],
) -> int:
    groups = collect_wave_groups(workbook)

    reset_sheet(workbook, SUMMARY_SHEET)
    summary = workbook.create_sheet(SUMMARY_SHEET)
    summary.append(SUMMARY_HEADERS)

    iso_year, iso_week, _ = kpi_day.isocalendar()
    report_date = format_date(kpi_day)
    report_week = f"{iso_year}-W{iso_week:02d}"
    report_month = kpi_day.replace(day=1).strftime("%Y/%m/%d")

    for wave_no, group in groups.items():
        qty_value = ""
        if group["has_qty"]:
            qty_value = int(group["qty"]) if group["qty"].is_integer() else round(group["qty"], 2)

        summary.append(
            [
                report_date,
                report_week,
                report_month,
                picker_names.get(wave_no) or fallback_name,
                wave_no,
                qty_value,
                group["count"],
                format_datetime(group["min_pick"]),
                format_datetime(group["max_pick"]),
            ]
        )

    style_header(summary)
    resize_columns(summary)
    return len(groups)


def remove_unwanted_sheets(workbook: Any) -> None:
    if MAIN_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"Excel cannot find main sheet: {MAIN_SHEET}")

    for sheet in list(workbook.worksheets):
        if sheet.title != MAIN_SHEET:
            workbook.remove(sheet)


def default_output_path(input_path: Path) -> Path:
    output_dir = input_path.parent / "output"
    if input_path.parent == Path.cwd():
        output_dir = Path.cwd() / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"{input_path.stem}_kpi_cleaned.xlsx"


def clean_workbook(input_path: Path, output_path: Path, kpi_day: date, args: argparse.Namespace) -> None:
    workbook = load_workbook(input_path)
    remove_unwanted_sheets(workbook)
    kept_rows, removed_rows = build_filtered_sheet(workbook, kpi_day)

    picker_names: dict[str, str] = {}
    if args.fetch_picker_names:
        groups = collect_wave_groups(workbook)
        default_start, default_end = picker_search_window(kpi_day)
        start_time = args.picker_start_time or default_start.strftime("%Y-%m-%d %H:%M:%S")
        end_time = args.picker_end_time or default_end.strftime("%Y-%m-%d %H:%M:%S")
        picker_names = resolve_picker_names(
            wave_nos=list(groups.keys()),
            storage_state=args.storage_state,
            start_time=start_time,
            end_time=end_time,
            page_size=args.picker_page_size,
            sample_limit=args.picker_sample_limit,
            wave_limit=args.picker_wave_limit,
            sleep_seconds=args.picker_sleep,
        )

    wave_count = build_summary_sheet(workbook, kpi_day, args.name, picker_names)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    start_at, end_at = kpi_window(kpi_day)
    exclude_start, exclude_end = pick_time_exclusion_window(kpi_day)
    print(f"KPI date: {kpi_day:%Y-%m-%d}")
    print(f"KPI export window: {start_at:%Y-%m-%d %H:%M:%S} ~ {end_at:%Y-%m-%d %H:%M:%S}")
    print(f"pick-time exclusion window: {exclude_start:%Y-%m-%d %H:%M:%S} ~ {exclude_end:%Y-%m-%d %H:%M:%S}")
    print(f"kept rows: {kept_rows}")
    print(f"removed rows: {removed_rows}")
    print(f"wave summary rows: {wave_count}")
    if args.fetch_picker_names:
        print(f"picker names found: {sum(1 for value in picker_names.values() if value)}")
    print(f"output file: {output_path.resolve()}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Clean OMP delivery export Excel and build KPI wave summary.")
    parser.add_argument("input", help="Raw exported Excel file path")
    parser.add_argument("--kpi-date", required=True, help="KPI date, YYYY-MM-DD, for example 2026-06-02")
    parser.add_argument("--name", default="", help="Fallback name when picker name cannot be fetched")
    parser.add_argument("--output", default="", help="Output Excel path, default writes to output directory")
    parser.add_argument("--fetch-picker-names", action="store_true", help="Fetch picker name from OMP progress logs")
    parser.add_argument("--storage-state", default="", help="Optional Playwright storage_state JSON")
    parser.add_argument("--picker-start-time", default="", help="Override wave search start time")
    parser.add_argument("--picker-end-time", default="", help="Override wave search end time")
    parser.add_argument("--picker-page-size", type=int, default=20)
    parser.add_argument("--picker-sample-limit", type=int, default=3)
    parser.add_argument("--picker-wave-limit", type=int, default=0, help="Only fetch the first N wave picker names; 0 means all")
    parser.add_argument("--picker-sleep", type=float, default=0.0)
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    output_path = Path(args.output).resolve() if args.output else default_output_path(input_path).resolve()
    clean_workbook(input_path, output_path, parse_kpi_date(args.kpi_date), args)


if __name__ == "__main__":
    main()
